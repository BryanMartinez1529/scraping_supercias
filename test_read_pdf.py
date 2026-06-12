import csv
import logging
import time
import os
import re
import requests
import pdfplumber
from datetime import datetime
from io import BytesIO
from dotenv import load_dotenv
import psycopg2
import pytesseract
from PIL import Image
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

pytesseract.pytesseract.tesseract_cmd = '/opt/homebrew/bin/tesseract'

# ==========================================
# CONFIGURACIÓN
# ==========================================
URL_SEARCH_COMPANY = 'https://appscvsgen.supercias.gob.ec/consultaCompanias/societario/busquedaCompanias.jsf'
BASE_URL = 'https://appscvsgen.supercias.gob.ec'

ERROR_FILE = 'scraping-supercias-by-exp-driver-chrome-db-error.csv'
PROCESSED_FILE = 'scraping-supercias-by-exp-driver-chrome-db-processed.csv'
load_dotenv()

# ==========================================
# CONFIGURACIÓN BASE DE DATOS
# ==========================================
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

# ==========================================
# CONFIGURACIÓN DE LÓGICA DE REGISTROS (LOGS)
# ==========================================
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logging.captureWarnings(True)
logging.getLogger('urllib3').setLevel(logging.ERROR)


def clean_for_search(text):
    import unicodedata
    if not text:
        return ""
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    text = text.upper().strip()
    text = re.sub(r'[^A-Z0-9\s&]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def cargar_auditores_externos(ruta_excel):
    registro = {}
    if not os.path.exists(ruta_excel):
        logging.warning(f"Archivo de auditores no encontrado en {ruta_excel}. Se omitirá el match por Excel.")
        return registro
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta_excel, read_only=True)
        sheet = wb.active
        
        headers = [str(cell).strip().upper() for cell in next(sheet.iter_rows(max_row=1, values_only=True))]
        
        idx_rnae = headers.index('RNAE')
        idx_nombre = headers.index('NOMBRE')
        idx_ruc = headers.index('IDENTIFICACIÓN')
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(idx_rnae, idx_nombre, idx_ruc):
                continue
            rnae = str(row[idx_rnae]).strip() if row[idx_rnae] is not None else ""
            nombre = str(row[idx_nombre]).strip() if row[idx_nombre] is not None else ""
            ruc = str(row[idx_ruc]).strip() if row[idx_ruc] is not None else ""
            
            if not rnae and not nombre:
                continue
                
            nombre_limpio = nombre.replace("''", '"').replace("'", "").strip()
            
            search_terms = []
            term_full = clean_for_search(nombre_limpio)
            if term_full:
                search_terms.append(term_full)
                
            matches_quotes = re.findall(r"['\"]+([^'\"]+)['\"]+", nombre)
            for mq in matches_quotes:
                term_q = clean_for_search(mq)
                if len(term_q) >= 4:
                    search_terms.append(term_q)
                    
            term_root = term_full
            suffixes = [
                r'\bCIA\b', r'\bLTDA\b', r'\bS\b', r'\bA\b', r'\bSAS\b',
                r'\bLIMITADA\b', r'\bGROUP\b', r'\bAUDITORES\b', r'\bCONSULTORES\b',
                r'\bASESORES\b', r'\bECUADOR\b', r'\bASSOCIATES\b', r'\bASOCIADOS\b',
                r'\bCL\b', r'\bAAE\b'
            ]
            for suff in suffixes:
                term_root = re.sub(suff, '', term_root).strip()
            term_root = re.sub(r'\s+', ' ', term_root).strip()
            if len(term_root) >= 5 and term_root not in search_terms:
                search_terms.append(term_root)
                
            info = {
                'rnae': rnae,
                'nombre_original': nombre,
                'nombre': nombre_limpio,
                'ruc': ruc,
                'search_terms': search_terms
            }
            
            if rnae:
                registro[rnae] = info
                
        logging.info(f"Cargados {len(registro)} auditores autorizados desde {ruta_excel}.")
    except Exception as e:
        logging.error(f"Error cargando base de auditores: {e}")
    return registro


# ==========================================
# FUNCIONES AUXILIARES - ORIGINALES
# ==========================================
def refresh_session(driver):
    driver.delete_all_cookies()
    driver.execute_script("window.localStorage.clear();")
    driver.get('about:blank')
    driver.get(URL_SEARCH_COMPANY)


def process_captcha(driver, prev_src=None):
    """
    Resuelve el captcha de búsqueda (un solo intento por llamada).
    Espera a que cambie la imagen si se provee prev_src.
    Retorna (captcha_bytes, current_src) o (None, None).
    """
    try:
        selector_img = ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td img"
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, selector_img))
        )
        captcha_element = driver.find_element(By.CSS_SELECTOR, selector_img)
        
        # En reintentos, esperar a que el src cambie
        if prev_src:
            try:
                WebDriverWait(driver, 8).until(
                    lambda d: d.find_element(By.CSS_SELECTOR, selector_img).get_attribute("src") != prev_src
                )
                captcha_element = driver.find_element(By.CSS_SELECTOR, selector_img)
                logging.info("Imagen del captcha de búsqueda recargada (src cambió)")
            except TimeoutException:
                logging.warning("El src de la imagen de búsqueda no cambió en 8s, intentando de todas formas...")

        WebDriverWait(driver, 10).until(lambda d: captcha_element.size['width'] > 0)
        current_src = captcha_element.get_attribute("src")

        captcha_bytes = captcha_element.screenshot_as_png
        image = Image.open(BytesIO(captcha_bytes)).convert('L')
        width, height = image.size
        image = image.resize((width * 2, height * 2), Image.Resampling.LANCZOS)
        
        # Usar whitelist para mejorar considerablemente la precisión (solo dígitos)
        text = pytesseract.image_to_string(image, config='--psm 7 -c tessedit_char_whitelist=0123456789').strip()
        text = re.sub(r'\D', '', text)
        logging.info(f"Captcha de búsqueda extraído (limpio): {text}")

        if not text:
            logging.warning("OCR no pudo extraer dígitos del captcha de búsqueda.")
            return None, current_src

        captcha_input = driver.find_element(By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td input")
        captcha_input.clear()
        captcha_input.send_keys(text)

        boton_buscar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-button-text.ui-c"))
        )
        boton_buscar.click()
        return captcha_bytes, current_src

    except Exception as e:
        logging.warning(f"Error procesando captcha de búsqueda: {e}")
        return None, None


def extract_company_data(driver, auto_complete_count):
    correos = []
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#panelDerecho>div")))

    provincia = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(4) tr:nth-child(1) td:nth-child(2) input")))
    texto_provincia = provincia.get_attribute("value")
    ciudad = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(4) tr:nth-child(1) td:last-child input")))
    texto_ciudad = ciudad.get_attribute("value")
    situacion = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(2) tr:nth-child(3) td:nth-child(5) textarea")))
    texto_situacion = situacion.get_attribute("value")

    logging.info(f"Provincia: {texto_provincia} | Ciudad: {texto_ciudad}")
    time.sleep(2)
    driver.execute_script("document.querySelector('#panelDerecho>div').scrollTop += 400;")

    boton_contactos = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(5)")))
    boton_contactos.click()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6)")))

    correo_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(2) td:last-child input")))
    correo_2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(3) td:last-child input")))
    telefono_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(1) td:nth-child(5) input")))
    telefono_2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(1) td:last-child input")))
    celular = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(1) td:nth-child(5) input")))

    correos.extend([
        texto_provincia.strip(), texto_ciudad.strip(),
        correo_1.get_attribute("value").strip(), correo_2.get_attribute("value").strip(),
        telefono_1.get_attribute("value").strip(), telefono_2.get_attribute("value").strip(),
        celular.get_attribute("value").strip(), texto_situacion.strip(),
        str(auto_complete_count)
    ])
    return correos


# ==========================================
# FUNCIONES NUEVAS - AUDITORÍA EXTERNA PDF
# ==========================================

def wait_for_loading(driver, timeout=90):
    """
    Espera a que los indicadores de 'Procesando' o 'statusDialog' desaparezcan de la pantalla.
    """
    start_time = time.time()
    logging.info("Esperando que finalice el estado 'Procesando'...")
    
    # Damos 1.5 segundos para que aparezca el modal de carga si la red/servidor tarda en responder
    time.sleep(1.5)
    
    while time.time() - start_time < timeout:
        still_loading = False
        try:
            # 1. Diálogo con ID statusDialog
            status_dialogs = driver.find_elements(By.ID, "statusDialog")
            for dialog in status_dialogs:
                if dialog.is_displayed():
                    still_loading = True
                    break
            
            # 2. Elementos con clase ui-blockui
            if not still_loading:
                blockers = driver.find_elements(By.CLASS_NAME, "ui-blockui")
                for blocker in blockers:
                    if blocker.is_displayed():
                        still_loading = True
                        break
            
            # 3. Cualquier diálogo visible con texto "procesando" o "cargando"
            if not still_loading:
                dialogs = driver.find_elements(By.CSS_SELECTOR, ".ui-dialog, .ui-dialog-content")
                for d in dialogs:
                    if d.is_displayed():
                        # Evitamos el diálogo del captcha
                        if d.get_attribute("id") == "dlgCaptcha":
                            continue
                        text = d.text.lower()
                        if "procesando" in text or "cargando" in text or "procesando..." in text:
                            still_loading = True
                            break
                            
            if not still_loading:
                break
                
        except Exception:
            # Si hay un error (ej. StaleElementReferenceException), asumimos que sigue cargando
            still_loading = True
            
        time.sleep(1)
        
    duration = round(time.time() - start_time, 2)
    logging.info(f"Espera de procesamiento finalizada en {duration} segundos.")


def process_captcha_modal(driver, max_intentos=10):
    """
    Resuelve el captcha del modal de Información Anual — hasta max_intentos.
    Retorna True si resolvió exitosamente o no se requería, False si falló tras los reintentos.
    """
    prev_src = None

    for intento in range(max_intentos):
        try:
            # Verificar si el modal de captcha está visible
            try:
                captcha_dialog = driver.find_element(By.CSS_SELECTOR, "#dlgCaptcha")
                if not captcha_dialog.is_displayed():
                    if intento == 0:
                        return True  # No había captcha
                    else:
                        logging.info("El captcha modal ya no está visible (resuelto exitosamente)")
                        return True
            except NoSuchElementException:
                if intento == 0:
                    return True
                else:
                    logging.info("El captcha modal desapareció (resuelto exitosamente)")
                    return True

            logging.info(f"Captcha modal detectado en flujo de Auditoría (intento {intento+1}/{max_intentos})")

            # Re-localizar el elemento de la imagen en cada intento
            captcha_img = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage"))
            )

            # En reintentos, esperar a que el src de la imagen cambie
            if intento > 0 and prev_src:
                logging.info("Esperando que la imagen del captcha modal se recargue...")
                try:
                    WebDriverWait(driver, 8).until(
                        lambda d: d.find_element(
                            By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage"
                        ).get_attribute("src") != prev_src
                    )
                    captcha_img = driver.find_element(By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage")
                    logging.info("Imagen del captcha modal recargada (src cambió)")
                except TimeoutException:
                    logging.warning("El src de la imagen no cambió en 8s, intentando de todas formas...")

            # Esperar a que la imagen tenga dimensiones reales y esté completa
            WebDriverWait(driver, 10).until(lambda d: captcha_img.size['width'] > 0)
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: d.execute_script(
                        "return arguments[0].naturalWidth > 0 && arguments[0].complete;",
                        captcha_img
                    )
                )
            except TimeoutException:
                logging.warning("La imagen del captcha no terminó de cargar por completo, esperando 2s extra...")
                time.sleep(2)

            # Guardar el src actual para comparar en el siguiente intento
            try:
                prev_src = captcha_img.get_attribute("src")
            except Exception:
                prev_src = None

            time.sleep(0.5)

            captcha_bytes = captcha_img.screenshot_as_png
            image = Image.open(BytesIO(captcha_bytes)).convert('L')
            width, height = image.size
            image = image.resize((width * 2, height * 2), Image.Resampling.LANCZOS)
            
            # Usar whitelist para mejorar considerablemente la precisión (solo dígitos)
            text = pytesseract.image_to_string(image, config='--psm 7 -c tessedit_char_whitelist=0123456789').strip()
            text = re.sub(r'\D', '', text)
            logging.info(f"Texto captcha modal extraído (limpio): {text}")

            if not text:
                logging.warning("OCR no pudo extraer dígitos del captcha modal, reintentando...")
                # Forzar recarga haciendo click en la imagen
                try:
                    captcha_img.click()
                    time.sleep(2)
                except Exception:
                    time.sleep(1)
                continue

            captcha_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#frmCaptcha\\:captcha"))
            )
            captcha_input.clear()
            captcha_input.send_keys(text)

            boton_verificar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#frmCaptcha\\:btnPresentarContenido"))
            )
            boton_verificar.click()
            logging.info("Captcha modal enviado")
            wait_for_loading(driver)

        except TimeoutException:
            if intento == 0:
                logging.info("No se detectó captcha modal")
                return True
            else:
                logging.warning("Timeout durante intento de resolver captcha modal")

    # Verificar por última vez si el modal sigue en pantalla
    try:
        captcha_dialog = driver.find_element(By.CSS_SELECTOR, "#dlgCaptcha")
        if captcha_dialog.is_displayed():
            logging.error("No se pudo superar el captcha modal tras los reintentos")
            return False
    except NoSuchElementException:
        pass

    return True


def navigate_to_informacion_anual(driver):
    try:
        menu_item = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']"))
        )
        menu_item.click()
        logging.info("Click en Información anual presentada")
        
        # Esperamos a que finalice el estado "Procesando"
        wait_for_loading(driver)

        # Esperamos a que aparezca el captcha modal o la tabla
        logging.info("Esperando captcha modal o tabla de información anual...")
        wait = WebDriverWait(driver, 45)
        
        class captcha_or_table_visible(object):
            def __call__(self, d):
                try:
                    dlg = d.find_element(By.CSS_SELECTOR, "#dlgCaptcha")
                    if dlg.is_displayed():
                        return "captcha"
                except Exception:
                    pass
                try:
                    tbl = d.find_element(By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']")
                    if tbl.is_displayed():
                        return "table"
                except Exception:
                    pass
                return False

        loaded_state = wait.until(captcha_or_table_visible())
        logging.info(f"Estado detectado: {loaded_state}")

        if loaded_state == "captcha":
            captcha_resuelto = process_captcha_modal(driver)
            if not captcha_resuelto:
                logging.warning("No se pudo resolver el captcha modal de Información Anual")
                return False
            # Esperamos de nuevo tras enviar el captcha
            wait_for_loading(driver)

        WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']"))
        )
        logging.info("Tabla de información anual cargada")
        return True

    except TimeoutException:
        logging.warning("No se pudo cargar la tabla de información anual")
        return False


def find_auditoria_externa_row(driver):
    page = 1
    while True:
        logging.info(f"Buscando 'Auditoria Externa' en página {page} de la tabla")
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']//tr"))
            )
            rows = driver.find_elements(By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']//tr[@data-ri]")

            for row in rows:
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) < 5:
                        continue
                    codigo_text = cells[1].text.strip()
                    nombre_text = cells[2].text.strip()

                    if codigo_text == '3.1.2' or 'Auditoria Externa' in nombre_text:
                        logging.info(f"Encontrada fila: código={codigo_text}, nombre={nombre_text}")
                        anio_fila = None
                        try:
                            anio_text = cells[0].text.strip()
                            anio_fila = int(anio_text) if anio_text.isdigit() else None
                            logging.info(f"Año extraído: {anio_fila}")
                        except Exception:
                            pass
                        pdf_link = cells[4].find_element(By.CSS_SELECTOR, "a.ui-commandlink")
                        return pdf_link, anio_fila

                except NoSuchElementException:
                    continue

        except TimeoutException:
            logging.warning("Timeout esperando filas de la tabla")
            return None, None

        try:
            next_btn = driver.find_element(
                By.XPATH,
                "//*[@id='frmInformacionCompanias:tblInformacionAnual_paginator_bottom']//a[contains(@class,'ui-paginator-next')]"
            )
            if 'ui-state-disabled' in next_btn.get_attribute('class'):
                logging.info("No hay más páginas. Auditoría externa no encontrada.")
                return None, None
            next_btn.click()
            time.sleep(2)
            page += 1
        except NoSuchElementException:
            logging.info("No se encontró paginador.")
            return None, None


def extract_pdf_url_from_modal(driver):
    """
    Extrae URL del PDF del modal — hasta 2 intentos.
    Maneja modal normal y modal de firma electrónica.
    """
    for intento in range(2):
        try:
            WebDriverWait(driver, 15).until(
                lambda d: (
                    _is_visible(d, "//*[@id='dlgPresentarDocumentoPdf']") or
                    _is_visible(d, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']")
                )
            )
        except TimeoutException:
            logging.warning(f"Ningún modal de PDF apareció (intento {intento+1})")
            if intento < 1:
                time.sleep(3)
                continue
            return None

        # ---- Caso 1: Modal normal con <object> ----
        try:
            dlg_normal = driver.find_element(By.XPATH, "//*[@id='dlgPresentarDocumentoPdf']")
            if dlg_normal.is_displayed():
                try:
                    obj_element = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located(
                            (By.XPATH, "//*[@id='panelInternoPresentarDocumentoPdf']//object[@type='application/pdf']")
                        )
                    )
                    pdf_data = obj_element.get_attribute("data")
                    if pdf_data:
                        pdf_url = pdf_data.split('?')[0]
                        if not pdf_url.startswith('http'):
                            pdf_url = BASE_URL + pdf_url
                        logging.info(f"URL del PDF extraída (modal normal): {pdf_url}")
                        return pdf_url
                except TimeoutException:
                    logging.warning(f"Modal normal sin <object> PDF (intento {intento+1})")
        except NoSuchElementException:
            pass

        # ---- Caso 2: Modal de firma electrónica ----
        try:
            dlg_firma = driver.find_element(By.XPATH, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']")
            if dlg_firma.is_displayed():
                logging.info("Modal de firma electrónica detectado")
                boton_aceptar = dlg_firma.find_element(
                    By.XPATH, ".//button[contains(@onclick,'window.open')]"
                )
                onclick = boton_aceptar.get_attribute("onclick")
                match = re.search(
                    r'window\.open\(.(/consultaCompanias/tmp/documento_[^.]+\.pdf)',
                    onclick
                )
                if match:
                    pdf_url = BASE_URL + match.group(1)
                    logging.info(f"URL PDF (firma electrónica): {pdf_url}")
                    try:
                        close_btn = dlg_firma.find_element(By.XPATH, ".//a[contains(@class,'ui-dialog-titlebar-close')]")
                        close_btn.click()
                        time.sleep(0.5)
                    except Exception:
                        pass
                    return pdf_url
        except NoSuchElementException:
            pass

        if intento < 1:
            logging.warning(f"No se extrajo URL del modal, reintentando en 3s...")
            time.sleep(3)

    logging.warning("No se pudo extraer URL del PDF después de 2 intentos")
    return None


def _is_visible(driver, xpath):
    try:
        el = driver.find_element(By.XPATH, xpath)
        return el.is_displayed()
    except Exception:
        return False


def download_pdf_in_memory(driver, pdf_url):
    # Método 1: fetch() desde el browser
    try:
        b64 = driver.execute_async_script(
            """
            var url = arguments[0];
            var done = arguments[1];
            fetch(url, {credentials: 'include'})
                .then(r => r.arrayBuffer())
                .then(buf => {
                    var bytes = new Uint8Array(buf);
                    var binary = '';
                    for (var i = 0; i < bytes.byteLength; i++) {
                        binary += String.fromCharCode(bytes[i]);
                    }
                    done(btoa(binary));
                })
                .catch(e => done(null));
            """,
            pdf_url
        )
        if b64:
            import base64
            pdf_content = base64.b64decode(b64)
            if len(pdf_content) > 1000:
                logging.info(f"PDF descargado via browser fetch: {len(pdf_content)} bytes")
                return BytesIO(pdf_content)
    except Exception as e:
        logging.warning(f"fetch() browser falló: {e}")

    # Método 2: requests con cookies
    try:
        selenium_cookies = driver.get_cookies()
        session = requests.Session()
        for cookie in selenium_cookies:
            session.cookies.set(cookie['name'], cookie['value'])
        headers = {
            'User-Agent': driver.execute_script("return navigator.userAgent;"),
            'Referer': driver.current_url,
        }
        response = session.get(pdf_url, headers=headers, timeout=30)
        response.raise_for_status()
        logging.info(f"PDF descargado via requests: {len(response.content)} bytes")
        return BytesIO(response.content)
    except Exception as e:
        logging.error(f"Error descargando PDF: {e}")
        return None


def extract_firma_auditora_from_pdf(pdf_bytes, base_auditores=None):
    """
    Extrae firma auditora y socio firmante del PDF de Auditoria Externa.

    Estrategias (en orden):
      1. pdfplumber en páginas 1-7 (texto digital)
      2. Footer fitz en páginas 3-6 (captura firma/socio del pie de página)
      3. OCR zona inferior páginas 5-8 (PDFs escaneados)

    Retorna: "PKFECUADOR & CO. C.L / Manuel García (SC-RNAE-002)"
             "Etl-Ec Auditores S..A / Nancy Proaño (Reg. 680)"
             "Ing. Paola Zamora C (SC-RNAE-1562)"
    """
    texto_completo_pdf = ""
    # Intentamos extraer todo el texto digital usando PyMuPDF (muy rápido)
    try:
        import fitz
        pdf_bytes.seek(0)
        doc = fitz.open(stream=pdf_bytes.read(), filetype="pdf")
        for num in range(min(8, doc.page_count)):
            page_text = doc[num].get_text()
            if page_text:
                texto_completo_pdf += " " + page_text
        doc.close()
    except Exception as e:
        logging.warning(f"Error al extraer texto preliminar con PyMuPDF: {e}")

    TITULOS_A_IGNORAR = [
        'dictamen de los auditores', 'informe de auditor', 'informe del auditor',
        'informe de los auditores', 'auditores independientes',
        'auditoria externa', 'estados financieros', 'parte i', 'parte ii',
        'externoindependiente', 'externo independiente', 'informe sobre',
        'cuenca, marzo', 'cuenca, ecuador', 'samborondon, ecuador',
        'quito, ecuador', 'quito abril', 'quito enero', 'quito febrero',
        'quito marzo', 'quito mayo', 'quito junio',
        'cumplimiento tributario', 'de cumplimiento', 'responsabilidades',
        'base de la opinion', 'incertidumbre material', 'opinion:',
        'hemos auditado', 'hemos llevado', 'a los accionistas', 'a los señores'
    ]
    PREFIJOS_PROF = ['ing.', 'dr.', 'dra.', 'lcdo.', 'lcda.', 'cpa ', 'eco.', 'mba.', 'abg.']
    PATRONES_EMPRESA = [
        'pkf', 'deloitte', 'kpmg', 'ernst & young', 'ey ', 'grant thornton',
        'bdo ', 'moore ', 'crowe', 'mazars', 'nexia', 'baker tilly',
        'etl ', 'etl-ec', 'etl global', 'global auditum',
        'hansen holm', 'price waterhouse', 'pwc',
        '& co.', '& co. c.l.', '& asociados', 'auditores y consultores',
        'firma auditora', 'cia. ltda', 'cia.ltda',
        'auditores s.', 'auditores externos',
    ]

    def es_titulo(linea):
        return any(t in linea.lower() for t in TITULOS_A_IGNORAR)

    def es_ruido_ocr(linea):
        palabras = linea.split()
        if not palabras: return True
        cortas = sum(1 for p in palabras if len(re.sub(r'[^a-zA-Z]', '', p)) <= 2)
        return len(palabras) > 5 and cortas / len(palabras) > 0.6

    def limpiar_ocr(linea):
        linea = re.sub(r'^[|\[\]!l1I]+\s*', '', linea).strip()
        return re.sub(r'[|_]{2,}', '', linea).strip()

    def es_cargo_o_ruido(nombre_l):
        cargos = ['gerente', 'general', 'contador', 'contadora', 'presidente', 'administrador', 'representante', 'legal', 'secretario', 'director', 'comisario']
        palabras = nombre_l.split()
        if not palabras:
            return True
        if all(p in cargos for p in palabras):
            return True
        if len(nombre_l) < 4:
            return True
        return False

    def normalizar_nombre(nombre):
        nombre = re.sub(r'^[nN][gG]\.', 'Ing.', nombre)
        nombre = re.sub(r'^\([nN][gG][,\.]', 'Ing.', nombre)
        nombre = re.sub(r'^\|\s*aola', 'Paola', nombre)
        nombre = re.sub(r'\s*\|.*$', '', nombre)
        return nombre.rstrip('.|,').strip()

    def extraer_rnae_numero(texto):
        m = re.search(r'SC-?RN\s?AE[-\s]*(\d+)', texto, re.IGNORECASE)
        if m:
            return f"SC-RNAE-{m.group(1)}"
        m2 = re.search(r'Registro\s+No\.?\s*(\d+)', texto, re.IGNORECASE)
        if m2:
            return f"Reg. {m2.group(1)}"
        return None

    def puntaje_candidato(nombre):
        nombre_norm = normalizar_nombre(nombre)
        nombre_l = nombre_norm.lower()
        puntaje = 0
        if any(p in nombre_l for p in PATRONES_EMPRESA): puntaje += 25
        if any(nombre_l.startswith(p) for p in PREFIJOS_PROF): puntaje += 20
        if 5 <= len(nombre_norm) <= 80: puntaje += 5
        if len(nombre_norm.split()) >= 2: puntaje += 5
        raros = sum(1 for c in nombre_norm if not (c.isalpha() or c in ' .-,°()&'))
        puntaje -= raros * 3
        palabras_1 = sum(1 for p in nombre_norm.split() if len(re.sub(r'[^a-zA-Z]', '', p)) == 1)
        puntaje -= palabras_1 * 5
        return puntaje, nombre_norm

    def buscar_en_lineas(lineas, solo_primeras_lineas=False):
        resultado = {'empresa': [], 'persona': [], 'socio': None, 'rnae': None}
        rango = range(min(5, len(lineas))) if solo_primeras_lineas else range(len(lineas))

        for i in rango:
            linea = lineas[i]
            lc = limpiar_ocr(linea.strip())
            lc_norm = normalizar_nombre(lc)
            lc_l = lc_norm.lower()

            if es_titulo(lc) or es_ruido_ocr(lc):
                continue

            # Firma empresa
            if any(p in lc_l for p in PATRONES_EMPRESA):
                if 3 < len(lc) <= 80 and len(lc.split()) <= 8:
                    resultado['empresa'].append(lc_norm)

            # Línea con RNAE / Registro
            es_registro_auditor = (
                'rnae' in lc_l or 'sc-rnae' in lc_l or
                'registro no.' in lc_l or 'registro no. sc' in lc_l or
                ('auditor scv' in lc_l and 'rnae' in lc_l) or
                ('licencia no.' in lc_l and 'registro no.' in lc_l)
            )
            if es_registro_auditor:
                rnae = extraer_rnae_numero(lc)
                if rnae and not resultado['rnae']:
                    resultado['rnae'] = rnae

                # Socio en la misma línea (después del RNAE)
                match = re.search(r'(?:SC-?RN\s+AE|SC-?RNAE)-?\s*\d+\s+(.+)', lc, re.IGNORECASE)
                if match:
                    socio_inline = match.group(1).strip().rstrip('.,')
                    if len(socio_inline) > 5 and len(socio_inline.split()) >= 2:
                        resultado['socio'] = socio_inline

                # Socio en líneas anteriores (sin filtro es_titulo para no perder "Quito abril... Nancy")
                if not match and i > 0:
                    for k in range(max(0, i-3), i):
                        linea_k = limpiar_ocr(lineas[k].strip())
                        nm = re.search(
                            r'\d{4}\s+([A-ZÁÉÍÓÚÑ][a-záéíóúñA-ZÁÉÍÓÚÑ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñA-ZÁÉÍÓÚÑ]+)+)',
                            linea_k
                        )
                        if nm:
                            nombre = nm.group(1).strip()
                            if len(nombre) > 5 and len(nombre.split()) >= 2:
                                resultado['socio'] = nombre
                                break

                if not resultado['socio']:
                    for j in range(max(0, i-4), i):
                        c = limpiar_ocr(lineas[j].strip())
                        c_norm = normalizar_nombre(c)
                        c_l = c_norm.lower()
                        if (len(c_norm) > 6 and not es_titulo(c_norm)
                                and not re.match(r'^[\d\s,\.°]+$', c_norm)
                                and not es_ruido_ocr(c_norm)
                                and not es_cargo_o_ruido(c_l)
                                and 'cuenca' not in c_l
                                and 'ecuador' not in c_l
                                and 'samborondon' not in c_l
                                and 'registro' not in c_l
                                and len(c_norm.split()) >= 2):
                            resultado['persona'].append(c_norm)
                            resultado['socio'] = resultado['socio'] or c_norm

            # Título profesional cerca de auditor/RNAE
            if any(lc_l.startswith(p) for p in PREFIJOS_PROF):
                ctx = ' '.join(lineas[max(0, i-2):min(len(lineas), i+5)]).lower()
                if ('auditor' in ctx or 'rnae' in ctx or 'socio' in ctx or
                        'scvc' in ctx or 'scvg' in ctx or 'pkf' in ctx):
                    if len(lc_norm) > 8 and len(lc_norm.split()) >= 3:
                        resultado['persona'].append(lc_norm)

        return resultado

    def extraer_desde_footer_fitz(pdf_bytes_local, desde_pct=0.80):
        """
        Extrae texto del footer (últimos 20%) usando PyMuPDF.
        Complementa pdfplumber para casos como ETL donde firma/socio están en el pie.
        """
        res = {'empresa': [], 'socio': None, 'rnae': None}
        try:
            import fitz
            pdf_bytes_local.seek(0)
            doc_f = fitz.open(stream=pdf_bytes_local.read(), filetype="pdf")
            total = doc_f.page_count
            paginas = list(range(2, min(6, total))) + list(range(0, min(2, total)))

            for num in paginas:
                page = doc_f[num]
                rect = page.rect
                zona = fitz.Rect(0, rect.height * desde_pct, rect.width, rect.height)
                texto = page.get_text("text", clip=zona).strip()
                if not texto:
                    continue

                lineas = [l.strip() for l in texto.split('\n') if l.strip()]
                logging.info(f"Footer fitz pag {num+1}: {lineas[:5]}")

                for i, linea in enumerate(lineas):
                    lc = limpiar_ocr(linea)
                    lc_l = lc.lower()

                    # Empresa en footer
                    if any(p in lc_l for p in PATRONES_EMPRESA):
                        if 3 < len(lc) <= 80 and len(lc.split()) <= 8 and not es_titulo(lc):
                            res['empresa'].append(lc)

                    # Registro/RNAE → buscar socio en líneas previas
                    es_reg = ('registro no.' in lc_l or 'rnae' in lc_l or
                              'sc-rnae' in lc_l or
                              ('licencia no.' in lc_l and 'registro' in lc_l))
                    if es_reg:
                        m = re.search(r'SC-?RN\s?AE[-\s]*(\d+)', lc, re.IGNORECASE)
                        if m and not res['rnae']:
                            res['rnae'] = f"SC-RNAE-{m.group(1)}"
                        m2 = re.search(r'Registro\s+No\.?\s*(\d+)', lc, re.IGNORECASE)
                        if m2 and not res['rnae']:
                            res['rnae'] = f"Reg. {m2.group(1)}"

                        if not res['socio']:
                            for k in range(max(0, i-5), i):
                                c = limpiar_ocr(lineas[k])
                                # "Ciudad mes año Nombre" → extraer nombre al final
                                nm = re.search(
                                    r'\d{4}\s+([A-ZÁÉÍÓÚÑ][a-záéíóúñA-ZÁÉÍÓÚÑ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñA-ZÁÉÍÓÚÑ]+)+)',
                                    c
                                )
                                if nm:
                                    nombre = nm.group(1).strip()
                                    if len(nombre) > 5 and len(nombre.split()) >= 2:
                                        res['socio'] = nombre
                                        break
                                # Nombre solo: 2-4 palabras, sin dígitos, no es empresa
                                elif (2 <= len(c.split()) <= 4
                                      and not re.search(r'\d', c)
                                      and not es_titulo(c)
                                      and not es_cargo_o_ruido(c.lower())
                                      and len(c) > 6
                                      and c[0].isupper()
                                      and not any(p in c.lower() for p in PATRONES_EMPRESA)):
                                    res['socio'] = c
                                    break

            doc_f.close()
        except Exception as e:
            logging.warning(f"extraer_desde_footer_fitz error: {e}")
        return res

    def ocr_zona(doc_fitz, num_pagina, desde=0.60):
        try:
            import fitz
            page = doc_fitz[num_pagina]
            mat = fitz.Matrix(3, 3)
            pix = page.get_pixmap(matrix=mat)
            image = Image.open(BytesIO(pix.tobytes("png"))).convert('L')
            w, h = image.size
            zona = image.crop((0, int(h * desde), w, h))
            return pytesseract.image_to_string(zona, config='--psm 6')
        except Exception as e:
            logging.warning(f"Error OCR zona pag {num_pagina+1}: {e}")
            return ""

    # ==========================================
    # ACUMULADORES
    # ==========================================
    todas_empresas = []
    todos_socios = []
    todos_rnae = []
    todos_personas = []

    # ---- INTENTO 1: pdfplumber páginas 1-7 (texto digital) ----
    try:
        pdf_bytes.seek(0)
        with pdfplumber.open(pdf_bytes) as pdf:
            total_paginas = len(pdf.pages)
            logging.info(f"PDF tiene {total_paginas} paginas")

            paginas_texto = list(range(min(5, total_paginas)))
            paginas_texto += [p for p in range(4, min(8, total_paginas)) if p not in paginas_texto]

            for num in paginas_texto:
                texto = pdf.pages[num].extract_text() or ""
                if not texto.strip(): continue
                texto_completo_pdf += " " + texto
                lineas = [l.strip() for l in texto.split('\n') if l.strip()]
                solo_inicio = num < 2
                r = buscar_en_lineas(lineas, solo_primeras_lineas=solo_inicio)
                if not r['empresa'] and not r['socio'] and solo_inicio:
                    r = buscar_en_lineas(lineas)
                todas_empresas.extend([(e, num+1) for e in r['empresa']])
                todos_personas.extend([(p, num+1) for p in r['persona']])
                if r['socio']: todos_socios.append(r['socio'])
                if r['rnae']: todos_rnae.append(r['rnae'])
    except Exception as e:
        logging.warning(f"Error pdfplumber: {e}")

    # ---- INTENTO 1b: Footer fitz (complementa pdfplumber) ----
    try:
        pdf_bytes.seek(0)
        pdf_bytes_copy = BytesIO(pdf_bytes.read())
        r_footer = extraer_desde_footer_fitz(pdf_bytes_copy)
        if r_footer['empresa']:
            todas_empresas.extend([(e, 0) for e in r_footer['empresa']])
            logging.info(f"Footer fitz empresa: {r_footer['empresa']}")
        if r_footer['socio'] and r_footer['socio'] not in todos_socios:
            todos_socios.append(r_footer['socio'])
            logging.info(f"Footer fitz socio: {r_footer['socio']}")
        if r_footer['rnae'] and not todos_rnae:
            todos_rnae.append(r_footer['rnae'])
            logging.info(f"Footer fitz rnae: {r_footer['rnae']}")
    except Exception as e:
        logging.warning(f"Footer fitz error: {e}")

    # ---- INTENTO 2: OCR zona inferior páginas 5-8 (PDF escaneado) ----
    if not todas_empresas and not todos_personas:
        try:
            import fitz
            pdf_bytes.seek(0)
            doc_ocr = fitz.open(stream=pdf_bytes.read(), filetype="pdf")
            total_paginas = doc_ocr.page_count
            paginas = (list(range(4, min(8, total_paginas))) +
                       list(range(1, min(4, total_paginas))))
            for num in paginas:
                texto = ocr_zona(doc_ocr, num, desde=0.60)
                texto_completo_pdf += " " + texto
                lineas = [l.strip() for l in texto.split('\n') if l.strip()]
                r = buscar_en_lineas(lineas)
                todas_empresas.extend([(e, num+1) for e in r['empresa']])
                todos_personas.extend([(p, num+1) for p in r['persona']])
                if r['socio']: todos_socios.append(r['socio'])
                if r['rnae']: todos_rnae.append(r['rnae'])
            doc_ocr.close()
        except ImportError:
            logging.warning("PyMuPDF no disponible")
        except Exception as e:
            logging.warning(f"Error OCR: {e}")

    # ---- Elegir mejor empresa y persona por puntaje ----
    mejor_empresa = None
    mejor_persona = None
    mejor_rnae = todos_rnae[0] if todos_rnae else None

    if todas_empresas:
        evaluados = [(puntaje_candidato(n)[0], puntaje_candidato(n)[1], 0) for n, pg in todas_empresas]
        evaluados.sort(reverse=True)
        mejor_empresa = evaluados[0][1]
        logging.info(f"Empresa auditora (candidata): '{mejor_empresa}'")

    if todos_personas:
        evaluados = [(puntaje_candidato(n)[0], puntaje_candidato(n)[1], 0) for n, _ in todos_personas]
        evaluados.sort(reverse=True)
        mejor_persona = evaluados[0][1]
        logging.info(f"Persona auditora (candidata): '{mejor_persona}'")

    socio_final = None
    if todos_socios:
        socios_eval = [(puntaje_candidato(s)[0], puntaje_candidato(s)[1]) for s in todos_socios]
        socios_eval.sort(reverse=True)
        socio_final = socios_eval[0][1]
        logging.info(f"Socio seleccionado (candidato): '{socio_final}'")

    # ---- MATCH CON LA BASE DE DATOS DE EXCEL ----
    matched_auditor = None
    if base_auditores:
        texto_pdf_limpio = clean_for_search(texto_completo_pdf)
        
        # 1. Intentar buscar por RNAE
        rnae_matches = re.findall(r'(?:SC\s?RN\s?AE|RNAE|REGISTRO|REG|REGISTRO\s?NO|REG\s?NO)[-\s]*(\d+)', texto_pdf_limpio)
        for r_num in rnae_matches:
            r_num_clean = str(int(r_num)) if r_num.isdigit() else r_num
            if r_num_clean in base_auditores:
                matched_auditor = base_auditores[r_num_clean]
                logging.info(f"Excel Match por RNAE '{r_num_clean}': {matched_auditor['nombre']}")
                break
                
        # 2. Si no hay match por RNAE, buscar por nombre
        if not matched_auditor:
            matches_by_name = []
            for r_num, info in base_auditores.items():
                for term in info['search_terms']:
                    if term in texto_pdf_limpio:
                        matches_by_name.append((len(term), info))
            if matches_by_name:
                matches_by_name.sort(key=lambda x: x[0], reverse=True)
                matched_auditor = matches_by_name[0][1]
                logging.info(f"Excel Match por Nombre: {matched_auditor['nombre']} (RNAE: {matched_auditor['rnae']})")

    # ---- Aplicar el match de la base de datos si existe ----
    if matched_auditor:
        es_empresa = any(w in matched_auditor['nombre'].upper() for w in ["CIA", "LTDA", "S.A", "SAS", "GROUP", "AUDITORES", "CONSULTORES", "ASESORES"])
        mejor_rnae = matched_auditor['rnae']
        if es_empresa:
            mejor_empresa = matched_auditor['nombre']
        else:
            mejor_persona = matched_auditor['nombre']
            mejor_empresa = None
            socio_final = None # No hace falta socio para persona natural

    # ---- Filtrar cargo o ruido del socio y persona ----
    if socio_final and es_cargo_o_ruido(socio_final.lower()):
        socio_final = None
    if mejor_persona and es_cargo_o_ruido(mejor_persona.lower()):
        mejor_persona = None

    # ---- Filtrar socio si es igual al nombre de la empresa ----
    if socio_final and matched_auditor:
        clean_socio = clean_for_search(socio_final)
        clean_company = clean_for_search(matched_auditor['nombre'])
        if clean_socio in clean_company or clean_company in clean_socio:
            socio_final = None

    if mejor_persona and socio_final:
        p_socio, _ = puntaje_candidato(socio_final)
        p_persona, _ = puntaje_candidato(mejor_persona)
        if p_persona > p_socio:
            socio_final = mejor_persona
    elif mejor_persona and not socio_final:
        socio_final = mejor_persona

    # Normalizar formato de mejor_rnae
    if mejor_rnae:
        mejor_rnae_str = str(mejor_rnae)
        if not mejor_rnae_str.startswith("SC-RNAE") and not mejor_rnae_str.startswith("Reg."):
            mejor_rnae = f"SC-RNAE-{mejor_rnae_str}"

    # ---- Construir resultado final ----
    partes = []
    if mejor_empresa:
        partes.append(mejor_empresa)
    nombre_auditor = socio_final or mejor_persona
    if nombre_auditor:
        if mejor_rnae:
            partes.append(f"{nombre_auditor} ({mejor_rnae})")
        else:
            partes.append(nombre_auditor)
    elif mejor_rnae and not mejor_empresa:
        partes.append(f"RNAE {mejor_rnae}")

    if partes:
        resultado_final = ' / '.join(partes)
        logging.info(f"Resultado final: '{resultado_final}'")
        return resultado_final

    logging.info("No se encontro firma auditora en el PDF")
    return None


def close_pdf_modal(driver):
    try:
        close_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='dlgPresentarDocumentoPdf']//a[contains(@class,'ui-dialog-titlebar-close')]"))
        )
        close_btn.click()
        time.sleep(1)
        logging.info("Modal PDF cerrado")
    except TimeoutException:
        pass

    try:
        close_btn = driver.find_element(
            By.XPATH, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']//a[contains(@class,'ui-dialog-titlebar-close')]"
        )
        if close_btn.is_displayed():
            close_btn.click()
            time.sleep(1)
    except NoSuchElementException:
        pass


def process_auditoria_externa(driver, expediente, base_auditores=None):
    resultado = {'firma': None, 'anio': None}
    try:
        if not navigate_to_informacion_anual(driver):
            logging.warning(f"[{expediente}] No se pudo navegar a Información Anual")
            return None

        pdf_link, anio_fila = find_auditoria_externa_row(driver)
        if not pdf_link:
            logging.info(f"[{expediente}] No tiene documento de Auditoría Externa — registrando S/A")
            return {'firma': 'S/A', 'anio': None}

        resultado['anio'] = anio_fila
        logging.info(f"[{expediente}] Año auditoría: {resultado['anio']}")

        logging.info(f"[{expediente}] Haciendo click en PDF de Auditoría Externa")
        pdf_link.click()
        
        # Esperamos a que finalice el estado "Procesando"
        wait_for_loading(driver)

        # Esperamos a que aparezca el captcha modal o el modal del PDF
        logging.info(f"[{expediente}] Esperando captcha modal o modal de PDF...")
        wait = WebDriverWait(driver, 45)
        
        class captcha_or_pdf_visible(object):
            def __call__(self, d):
                try:
                    dlg = d.find_element(By.CSS_SELECTOR, "#dlgCaptcha")
                    if dlg.is_displayed():
                        return "captcha"
                except Exception:
                    pass
                try:
                    if _is_visible(d, "//*[@id='dlgPresentarDocumentoPdf']") or \
                       _is_visible(d, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']"):
                        return "pdf"
                except Exception:
                    pass
                return False

        loaded_state = wait.until(captcha_or_pdf_visible())
        logging.info(f"[{expediente}] Estado detectado tras click en PDF: {loaded_state}")

        if loaded_state == "captcha":
            captcha_resuelto = process_captcha_modal(driver)
            if not captcha_resuelto:
                logging.warning(f"[{expediente}] No se pudo resolver el captcha del PDF")
                return None
            # Esperamos de nuevo tras resolver el captcha
            wait_for_loading(driver)

        pdf_url = extract_pdf_url_from_modal(driver)
        if not pdf_url:
            logging.warning(f"[{expediente}] No se pudo obtener URL del PDF")
            close_pdf_modal(driver)
            return None

        pdf_bytes = download_pdf_in_memory(driver, pdf_url)
        if not pdf_bytes:
            logging.warning(f"[{expediente}] Error descargando el PDF")
            close_pdf_modal(driver)
            return None

        firma = extract_firma_auditora_from_pdf(pdf_bytes, base_auditores)
        resultado['firma'] = firma
        logging.info(f"[{expediente}] Resultado: firma={firma}, anio={resultado['anio']}")

        close_pdf_modal(driver)
        return resultado

    except Exception as e:
        logging.error(f"[{expediente}] Error en process_auditoria_externa: {e}")
        try:
            close_pdf_modal(driver)
        except Exception:
            pass
        return None


def insertar_auditoria_en_bd(conn, id_compania, firma, anio):
    """Usa su propio cursor para no interferir con el cursor del loop principal."""
    if not firma:
        logging.info("Firma vacía, no se inserta en BD")
        return False
    cur = None
    try:
        cur = conn.cursor()
        # anio puede ser None (N/A) — usar IS NULL en ese caso
        if anio is None:
            cur.execute(
                "SELECT id FROM crm.com_companias_auditorias_previas WHERE id_compania = %s AND anio_auditoria IS NULL;",
                (id_compania,)
            )
        else:
            cur.execute(
                "SELECT id FROM crm.com_companias_auditorias_previas WHERE id_compania = %s AND anio_auditoria = %s;",
                (id_compania, anio)
            )
        existente = cur.fetchone()
        if existente:
            cur.execute(
                "UPDATE crm.com_companias_auditorias_previas SET nombre_firma = %s, fecha_creacion = now() WHERE id_compania = %s AND anio_auditoria = %s;",
                (firma, id_compania, anio)
            )
            logging.info(f"Registro actualizado: id={id_compania}, firma={firma}, anio={anio}")
        else:
            cur.execute(
                "INSERT INTO crm.com_companias_auditorias_previas (id_compania, nombre_firma, anio_auditoria, fecha_creacion) VALUES (%s, %s, %s, now());",
                (id_compania, firma, anio)
            )
            logging.info(f"Registro insertado: id={id_compania}, firma={firma}, anio={anio}")
        conn.commit()
        return True
    except Exception as e:
        logging.error(f"Error insertando auditoría en BD: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        return False
    finally:
        if cur:
            cur.close()


# ==========================================
# LÓGICA PRINCIPAL - ORIGINAL
# ==========================================
def search_company(compania, driver, new_connection):
    correos = []
    auto_complete_count = 0
    captcha_bytes = None

    try:
        if not new_connection:
            driver.get(URL_SEARCH_COMPANY)
        time.sleep(1)

        tipo_busqueda = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-selectoneradio.ui-widget td:nth-child(1)>div"))
        )
        tipo_busqueda.click()
        time.sleep(1)

        parametro_busqueda = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-autocomplete>input"))
        )
        parametro_busqueda.send_keys(compania)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all.ui-state-highlight"))
        )
        auto_complete_items = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all"))
        )
        auto_complete_count = len(auto_complete_items)

        auto_complete_item = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all.ui-state-highlight"))
        )
        auto_complete_item.click()

        data_result_elem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content tr:nth-child(1) td:nth-child(2) .ui-outputlabel.ui-widget"))
        )
        data_result = data_result_elem.text
        logging.info(f"Compañia encontrada: {data_result}")

        if int(data_result) != int(compania):
            logging.info("No se encontró el mismo expediente")
            refresh_session(driver)
            return None

        captcha_bytes, _ = process_captcha(driver)

    except StaleElementReferenceException as e:
        logging.info(f"Elemento stale: {e}")
    finally:
        if not captcha_bytes and auto_complete_count == 1:
            correos.extend(['N/D', 'N/D', 'N/D', 'N/D', 'N/D', 'N/D', 'N/D', 'N/D', str(auto_complete_count)])
            refresh_session(driver)
            return correos

    try:
        correos = extract_company_data(driver, auto_complete_count)
    except StaleElementReferenceException:
        pass
    finally:
        elements = driver.find_elements(By.CSS_SELECTOR, ".ui-messages-error-summary")
        if elements:
            refresh_session(driver)
        else:
            try:
                nueva_busqueda = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#frmNuevaConsulta>button"))
                )
                nueva_busqueda.click()
            except Exception:
                pass

    return correos


def process_database_records(error_file, processed_file, driver, new_connection, base_auditores=None):
    l_i = 1
    try:
        conn = psycopg2.connect(
            host=DB_HOST, port=DB_PORT, database=DB_NAME,
            user=DB_USER, password=DB_PASSWORD
        )
        cursor = conn.cursor()

        cursor.execute("""
            SELECT expediente, id
            FROM crm.adm_companias
            WHERE estado_proceso = 'T'
              AND fecha_actualizacion = '2026-05-26'
              AND activos >= 642000 and activos <= 3000000 and provincia in ('PICHINCHA', 'GUAYAS', 'AZUAY', 'LOJA', 'COTOPAXI', 'CHIMBORAZO', 'GUAYAQUIL', 'MANABI', 'SANTA ELENA', 'IMBABURA', 'EL ORO', 'TUNGURAHUA');
        """)
        records = cursor.fetchall()

    except Exception as e:
        logging.error(f"Error conectando a la base de datos: {e}")
        return

    try:
        with open(error_file, 'a', newline='') as csvfile_error, \
             open(processed_file, 'a', newline='') as csvfile_processed:

            error_writer = csv.writer(csvfile_error, delimiter=';')
            processed_writer = csv.writer(csvfile_processed, delimiter=';')

            for record in records:
                expediente = record[0]
                id_compania = record[1]
                row = [expediente]

                logging.info(f"Row: {l_i} | Company: {expediente} (id={id_compania})")

                success = False
                for intento in range(3):
                    try:
                        logging.info(f"[{expediente}] Navegando para extraccion de Auditoria Externa...")
                        _navegar_a_compania(driver, expediente)
                        auditoria = process_auditoria_externa(driver, expediente, base_auditores)

                        firma_insertada = False
                        if auditoria and auditoria.get('firma'):
                            firma_insertada = insertar_auditoria_en_bd(
                                conn,
                                id_compania=id_compania,
                                firma=auditoria['firma'],
                                anio=auditoria.get('anio')
                            )
                        else:
                            logging.info(f"[{expediente}] Sin firma auditora para registrar")

                        processed_writer.writerow(row)
                        success = True

                        if firma_insertada:
                            try:
                                cur_upd = conn.cursor()
                                cur_upd.execute("""
                                    UPDATE crm.adm_companias
                                    SET estado_proceso = 'P'
                                    WHERE id = %s;
                                """, (id_compania,))
                                conn.commit()
                                cur_upd.close()
                                logging.info(f"[{expediente}] estado_proceso → 'P'")
                            except Exception as e_upd:
                                logging.warning(f"[{expediente}] No se pudo actualizar estado_proceso: {e_upd}")
                                try:
                                    conn.rollback()
                                except Exception:
                                    pass
                        else:
                            logging.info(f"[{expediente}] estado_proceso NO actualizado (sin firma)")
                        break

                    except Exception as e:
                        logging.info(f"Error procesando {expediente} (intento {intento+1}): {e}")
                        time.sleep(2)

                if not success:
                    logging.info(f"Revisar {expediente}")
                    error_writer.writerow(row)

                l_i += 1

    finally:
        if cursor: cursor.close()
        if conn: conn.close()
        logging.info("Conexión BD cerrada.")


def _navegar_a_compania(driver, expediente, max_intentos_captcha=5):
    logging.info(f"Navegando a compañía {expediente}")
    driver.get(URL_SEARCH_COMPANY)
    time.sleep(2)

    tipo_busqueda = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-selectoneradio.ui-widget td:nth-child(1)>div"))
    )
    tipo_busqueda.click()
    time.sleep(1)

    parametro_busqueda = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-autocomplete>input"))
    )
    parametro_busqueda.send_keys(str(expediente))

    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all.ui-state-highlight"))
    )
    auto_complete_item = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all.ui-state-highlight"))
    )
    auto_complete_item.click()
    time.sleep(2)

    # Validar que el expediente sea el correcto antes de continuar
    data_result_elem = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content tr:nth-child(1) td:nth-child(2) .ui-outputlabel.ui-widget"))
    )
    data_result = data_result_elem.text.strip()
    if int(data_result) != int(expediente):
        raise ValueError(f"El expediente no coincide: buscado {expediente}, encontrado {data_result}")

    prev_src = None
    for intento in range(max_intentos_captcha):
        # 1. ¿Ya cargó el menú principal de la compañía?
        try:
            menu_present = driver.find_elements(By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']")
            if menu_present and menu_present[0].is_displayed():
                logging.info(f"Compañía {expediente} cargada con éxito")
                return True
        except Exception:
            pass

        # 2. ¿Sigue visible el captcha?
        try:
            captcha_img = driver.find_elements(By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td img")
            if not captcha_img or not captcha_img[0].is_displayed():
                # Esperar 2 segundos por si la respuesta al captcha exitoso está cargando lentamente
                time.sleep(2)
                menu_present = driver.find_elements(By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']")
                if menu_present and menu_present[0].is_displayed():
                    logging.info(f"Compañía {expediente} cargada con éxito (tras breve espera)")
                    return True
        except Exception as e:
            logging.warning(f"Error verificando presencia de captcha: {e}")

        # 3. Intentar resolver el captcha
        logging.info(f"Resolviendo captcha de búsqueda (intento {intento+1}/{max_intentos_captcha})...")
        captcha_bytes, current_src = process_captcha(driver, prev_src=prev_src)
        
        if current_src:
            prev_src = current_src

        # Esperar a que se procese la petición
        time.sleep(5)

        # 4. Verificar si hay errores (ej. captcha incorrecto)
        errors = driver.find_elements(By.CSS_SELECTOR, ".ui-messages-error-summary")
        if errors:
            logging.warning(f"Error detectado tras enviar captcha: {[e.text for e in errors]}. Reintentando...")
            continue

        # 5. Esperar a que el menú principal esté presente
        try:
            WebDriverWait(driver, 35).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']"))
            )
            logging.info(f"Compañía {expediente} cargada correctamente")
            return True
        except TimeoutException:
            logging.warning("El menú de la compañía no apareció tras 35 segundos. Verificando captcha...")

    raise TimeoutException(f"No se pudo cargar la compañía {expediente} tras {max_intentos_captcha} intentos de captcha.")


# ==========================================
# EJECUCIÓN
# ==========================================
def main():
    driver = None
    new_connection = False
    start_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        logging.info("Inicio del script.")
        options = Options()
        options.binary_location = '/Applications/Firefox.app/Contents/MacOS/firefox'
        # options.add_argument('--headless')
        driver = webdriver.Firefox(options=options)

        # Cargar base de datos de auditores externos desde Excel
        ruta_excel = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'auditores_externos.xlsx')
        base_auditores = cargar_auditores_externos(ruta_excel)

        try:
            process_database_records(ERROR_FILE, PROCESSED_FILE, driver, new_connection, base_auditores)
        except KeyboardInterrupt:
            logging.info("Proceso interrumpido por el usuario.")

        end_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{start_time_str} - {end_time_str}] Fin del script.")

    finally:
        if driver is not None:
            driver.quit()


if __name__ == "__main__":
    main()