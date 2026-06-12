import csv
import logging
import time
import os
import re
import requests
import pdfplumber
import json
from datetime import datetime
from io import BytesIO
from dotenv import load_dotenv
import psycopg2
import pytesseract
from PIL import Image
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

try:
    import ollama
except ImportError:
    ollama = None

load_dotenv()

# ==========================================
# CONFIGURACIÓN DE LÓGICA DE REGISTROS (LOGS)
# Debe ir ANTES de cualquier función que use logging al importar el módulo
# ==========================================
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    force=True
)
logging.captureWarnings(True)
logging.getLogger('urllib3').setLevel(logging.ERROR)

# ==========================================
# CONFIGURACIÓN DE BINARIOS Y FILTROS DESDE ENV
# ==========================================
TESSERACT_PATH = os.getenv("TESSERACT_PATH", "/opt/homebrew/bin/tesseract")
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

FIREFOX_BINARY_PATH = os.getenv("FIREFOX_BINARY_PATH", "/Applications/Firefox.app/Contents/MacOS/firefox")
FIREFOX_HEADLESS = os.getenv("FIREFOX_HEADLESS", "false").lower() in ("true", "1", "yes")

DB_MIN_ACTIVOS = int(os.getenv("DB_MIN_ACTIVOS", "642000"))
DB_FECHA_ACTUALIZACION = os.getenv("DB_FECHA_ACTUALIZACION", "2026-05-26")

AUDITORS_EXCEL_PATH = os.getenv("AUDITORS_EXCEL_PATH", "auditores_externos.xlsx")
AUDITORS_COLUMN_NAME = os.getenv("AUDITORS_COLUMN_NAME", "NOMBRE")

# ==========================================
# CONFIGURACIÓN DE OLLAMA
# ==========================================
USE_OLLAMA = os.getenv("USE_OLLAMA", "true").lower() in ("true", "1", "yes")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3.2:latest")
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://localhost:11434")


def _normalizar_texto(texto):
    """
    Normaliza el texto para facilitar la comparación:
    - Pasa a minúsculas
    - Remueve tildes y diacríticos
    - Remueve caracteres especiales y espacios múltiples
    """
    import unicodedata
    if not texto:
        return ""
    texto = str(texto).lower()
    # Remover tildes
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                     if unicodedata.category(c) != 'Mn')
    # Conservar solo caracteres alfanuméricos y espacios
    texto = re.sub(r'[^a-z0-9\s]', ' ', texto)
    # Colapsar espacios múltiples
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto


def cargar_lista_auditores():
    """
    Carga la lista de nombres de auditores desde el archivo Excel configurado.
    Retorna un conjunto de nombres normalizados y un mapa de nombre_normalizado -> nombre_original.
    """
    if not os.path.exists(AUDITORS_EXCEL_PATH):
        logging.warning(f"Archivo Excel de auditores '{AUDITORS_EXCEL_PATH}' no encontrado. Se usará coincidencia heurística estándar sin Excel.")
        return set(), {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(AUDITORS_EXCEL_PATH, read_only=True)
        sheet = wb.active
        
        # Encontrar la columna del nombre
        first_row = next(sheet.iter_rows(max_row=1))
        headers = [cell.value for cell in first_row]
        
        column_index = -1
        # Buscar coincidencia exacta
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == AUDITORS_COLUMN_NAME.lower():
                column_index = i
                break
                
        # Si no hay exacta, buscar coincidencia parcial
        if column_index == -1:
            for i, h in enumerate(headers):
                if h and AUDITORS_COLUMN_NAME.lower() in str(h).lower():
                    column_index = i
                    break
                    
        if column_index == -1:
            logging.error(f"Columna '{AUDITORS_COLUMN_NAME}' no encontrada en el Excel. Cabeceras disponibles: {headers}")
            return set(), {}

        auditores_normalizados = set()
        mapa_originales = {}

        # Leer filas
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) > column_index:
                val = row[column_index]
                if val:
                    nombre = str(val).strip()
                    if len(nombre) > 2:
                        norm = _normalizar_texto(nombre)
                        auditores_normalizados.add(norm)
                        mapa_originales[norm] = nombre

        logging.info(f"Se cargaron {len(auditores_normalizados)} auditores/firmas desde '{AUDITORS_EXCEL_PATH}'")
        return auditores_normalizados, mapa_originales

    except Exception as e:
        logging.error(f"Error leyendo el Excel de auditores: {e}")
        return set(), {}


def buscar_auditor_en_linea(linea, auditores_normalizados, mapa_originales, umbral=0.82):
    """
    Busca si algún auditor de la lista de Excel coincide exacta o difusamente con la línea de texto.
    """
    linea_norm = _normalizar_texto(linea)
    if not linea_norm:
        return None
        
    # 1. Coincidencia exacta de subcadena (más rápido)
    for aud_norm in auditores_normalizados:
        if aud_norm in linea_norm:
            return mapa_originales[aud_norm]

    # Coincidencia inversa (el extraído es una subcadena del nombre oficial largo de Excel)
    for aud_norm in auditores_normalizados:
        if len(linea_norm) >= 5 and linea_norm in aud_norm:
            return mapa_originales[aud_norm]
            
    # 2. Coincidencia difusa
    from difflib import SequenceMatcher
    for aud_norm in auditores_normalizados:
        len_aud = len(aud_norm)
        len_linea = len(linea_norm)
        
        # Comparación directa si las longitudes son similares
        if abs(len_linea - len_aud) <= 5:
            if SequenceMatcher(None, aud_norm, linea_norm).ratio() >= umbral:
                return mapa_originales[aud_norm]
        else:
            # Ventana deslizante para buscar subcadenas similares dentro de la línea
            for w_size in [len_aud - 1, len_aud, len_aud + 1, len_aud + 2]:
                if w_size > len_linea:
                    continue
                for i in range(0, len_linea - w_size + 1):
                    sub = linea_norm[i:i+w_size]
                    if SequenceMatcher(None, aud_norm, sub).ratio() >= umbral:
                        return mapa_originales[aud_norm]
                        
    return None

# Cargar la lista global de auditores
AUDITORES_SET, MAPA_ORIGINALES_AUD = cargar_lista_auditores()

# ==========================================
# CONFIGURACIÓN
# ==========================================
URL_SEARCH_COMPANY = 'https://appscvsgen.supercias.gob.ec/consultaCompanias/societario/busquedaCompanias.jsf'
BASE_URL = 'https://appscvsgen.supercias.gob.ec'

ERROR_FILE = 'scraping-supercias-by-exp-driver-chrome-db-error.csv'
PROCESSED_FILE = 'scraping-supercias-by-exp-driver-chrome-db-processed.csv'

# ==========================================
# CONFIGURACIÓN BASE DE DATOS
# ==========================================
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

# Nota: logging.basicConfig ya fue configurado al inicio del módulo
# para que los mensajes de cargar_lista_auditores() sean visibles.


# ==========================================
# FUNCIONES AUXILIARES - ORIGINALES
# ==========================================
def refresh_session(driver):
    """Refresca la página, limpia cookies y local storage."""
    driver.delete_all_cookies()
    driver.execute_script("window.localStorage.clear();")
    driver.get('about:blank')
    driver.get(URL_SEARCH_COMPANY)


def process_captcha(driver):
    """Espera, captura, procesa y resuelve el captcha de búsqueda."""
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td img"))
    )
    logging.info("Imagen del captcha encontrada")

    captcha_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td img"))
    )

    WebDriverWait(driver, 10).until(lambda d: captcha_element.size['width'] > 0)
    logging.info("Elemento del captcha cargado")

    captcha_bytes = captcha_element.screenshot_as_png
    logging.info("Imagen del captcha capturada en memoria")

    image = Image.open(BytesIO(captcha_bytes))
    image = image.convert('L')
    width, height = image.size
    image = image.resize((width * 2, height * 2), Image.Resampling.LANCZOS)

    text = pytesseract.image_to_string(image, config='--psm 7').strip()
    # Limpieza estricta de caracteres no numéricos
    text = re.sub(r'\D', '', text)
    logging.info(f"Texto del captcha extraído (limpio): {text}")

    captcha_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td input"))
    )
    captcha_input.clear()
    captcha_input.send_keys(text)
    logging.info("Captcha ingresado")

    boton_buscar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-button-text.ui-c"))
    )
    boton_buscar.click()
    logging.info("Boton buscar presionado")
    
    return captcha_bytes


def _buscar_y_cargar_compania(driver, expediente, max_intentos_captcha=3):
    """
    Realiza la búsqueda de una compañía por expediente y maneja la resolución
    del captcha de búsqueda con reintentos si es necesario.
    Retorna (success_bool, auto_complete_count)
    """
    expediente_str = str(expediente).strip()
    logging.info(f"Iniciando búsqueda de compañía: {expediente_str}")

    driver.get(URL_SEARCH_COMPANY)
    time.sleep(1.5)

    # 1. Seleccionar búsqueda por expediente
    tipo_busqueda = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-selectoneradio.ui-widget td:nth-child(1)>div"))
    )
    tipo_busqueda.click()
    time.sleep(1)

    # 2. Ingresar parámetro de búsqueda
    parametro_busqueda = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-autocomplete>input"))
    )
    parametro_busqueda.clear()
    parametro_busqueda.send_keys(expediente_str)
    logging.info(f"Expediente ingresado: {expediente_str}")

    # 3. Esperar autocompletar
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all.ui-state-highlight"))
    )
    auto_complete_items = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all"))
    )
    auto_complete_count = len(auto_complete_items)
    logging.info(f"Número de elementos de autocompletar: {auto_complete_count}")

    auto_complete_item = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-autocomplete-item.ui-autocomplete-list-item.ui-corner-all.ui-state-highlight"))
    )
    auto_complete_item.click()
    logging.info("Elemento de autocompletar seleccionado")
    time.sleep(2)

    # 4. Verificar que se encontró el expediente correcto
    data_result_elem = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content tr:nth-child(1) td:nth-child(2) .ui-outputlabel.ui-widget"))
    )
    data_result = data_result_elem.text.strip()
    logging.info(f"Compañía encontrada (expediente): {data_result}")

    if int(data_result) != int(expediente_str):
        logging.warning(f"No coincide el expediente encontrado ({data_result}) con el buscado ({expediente_str})")
        refresh_session(driver)
        return False, auto_complete_count

    # 5. Manejar Captcha con bucle de reintentos
    for intento in range(max_intentos_captcha):
        # Verificar si ya pasamos la pantalla de búsqueda (ej. cargó el menú)
        try:
            menu_present = driver.find_elements(By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']")
            if menu_present and menu_present[0].is_displayed():
                logging.info("Búsqueda exitosa, se cargó la compañía directamente (sesión previa activa o sin captcha)")
                return True, auto_complete_count
        except Exception:
            pass

        # Verificar si el captcha de búsqueda está visible
        try:
            captcha_img = driver.find_elements(By.CSS_SELECTOR, ".ui-panel-content.ui-widget-content>table>tbody>tr:nth-child(4)>td img")
            if not captcha_img or not captcha_img[0].is_displayed():
                logging.info("No se detectó imagen de captcha de búsqueda activa")
                time.sleep(1)
                menu_present = driver.find_elements(By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']")
                if menu_present and menu_present[0].is_displayed():
                    return True, auto_complete_count
                return False, auto_complete_count
        except Exception:
            return False, auto_complete_count

        logging.info(f"Resolviendo captcha de búsqueda (intento {intento+1}/{max_intentos_captcha})...")
        process_captcha(driver)
        time.sleep(4)

        # Verificar si hay mensajes de error en pantalla (como captcha incorrecto)
        errors = driver.find_elements(By.CSS_SELECTOR, ".ui-messages-error-summary")
        if errors:
            logging.warning(f"Error detectado al enviar captcha: {[e.text for e in errors]}. Reintentando...")
            continue
        
        # Si no hay errores, verificar si se cargó la página de la compañía
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']"))
            )
            logging.info("Captcha de búsqueda resuelto correctamente")
            return True, auto_complete_count
        except TimeoutException:
            logging.warning("No se cargó el menú tras resolver el captcha. Reintentando...")

    logging.error("No se pudo superar el captcha de búsqueda después de los intentos permitidos")
    return False, auto_complete_count


def extract_company_data(driver, auto_complete_count):
    """Extrae la información de la compañía desde el panel desplegable."""
    correos = []

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#panelDerecho>div")))

    provincia = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(4) tr:nth-child(1) td:nth-child(2) input")))
    texto_provincia = provincia.get_attribute("value")

    ciudad = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(4) tr:nth-child(1) td:last-child input")))
    texto_ciudad = ciudad.get_attribute("value")

    situacion = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(2) tr:nth-child(3) td:nth-child(5) textarea")))
    texto_situacion = situacion.get_attribute("value")

    logging.info(f"Provincia: {texto_provincia}")
    logging.info(f"Ciudad: {texto_ciudad}")
    logging.info(f"Situacion: {texto_situacion}")
    time.sleep(2)

    driver.execute_script("document.querySelector('#panelDerecho>div').scrollTop += 400;")
    logging.info("Scroll vertical realizado")

    boton_contactos = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(5)"))
    )
    boton_contactos.click()

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6)"))
    )

    correo_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(2) td:last-child input")))
    texto_correo_1 = correo_1.get_attribute("value")

    correo_2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(3) td:last-child input")))
    texto_correo_2 = correo_2.get_attribute("value")

    telefono_1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(1) td:nth-child(5) input")))
    texto_telefono_1 = telefono_1.get_attribute("value")

    telefono_2 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(1) td:last-child input")))
    texto_telefono_2 = telefono_2.get_attribute("value")

    celular = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".ui-accordion.ui-widget.ui-helper-reset.ui-hidden-container>div:nth-child(6) tr:nth-child(1) td:nth-child(5) input")))
    texto_celular = celular.get_attribute("value")

    correos.extend([
        texto_provincia.strip(),
        texto_ciudad.strip(),
        texto_correo_1.strip(),
        texto_correo_2.strip(),
        texto_telefono_1.strip(),
        texto_telefono_2.strip(),
        texto_celular.strip(),
        texto_situacion.strip(),
        str(auto_complete_count)
    ])

    return correos


# ==========================================
# FUNCIONES NUEVAS - AUDITORÍA EXTERNA PDF
# ==========================================

def process_captcha_modal(driver, max_intentos=5):
    """
    Resuelve el captcha cuando aparece dentro del flujo de Información Anual.
    Usa el selector específico del modal dlgCaptcha de esa sección.
    Retorna True si resolvió exitosamente o no se requería, False si falló tras los reintentos.
    """
    prev_src = None  # Guardamos el src anterior para detectar cambio de imagen

    for intento in range(max_intentos):
        try:
            # Verificar si el modal de captcha está visible
            try:
                captcha_dialog = driver.find_element(By.CSS_SELECTOR, "#dlgCaptcha")
                if not captcha_dialog.is_displayed():
                    if intento == 0:
                        return True  # No había captcha
                    else:
                        logging.info("El captcha modal ya no está visible (resuelto exitosamente)")
                        return True
            except NoSuchElementException:
                if intento == 0:
                    return True
                else:
                    logging.info("El captcha modal desapareció (resuelto exitosamente)")
                    return True

            logging.info(f"Captcha modal detectado en flujo de Auditoría (intento {intento+1}/{max_intentos})")

            # Re-localizar el elemento de la imagen en cada intento (evita stale element)
            captcha_img = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage"))
            )

            # En reintentos, esperar a que el src de la imagen cambie (nueva imagen cargada)
            if intento > 0 and prev_src:
                logging.info("Esperando que la imagen del captcha modal se recargue...")
                try:
                    WebDriverWait(driver, 8).until(
                        lambda d: d.find_element(
                            By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage"
                        ).get_attribute("src") != prev_src
                    )
                    # Re-localizar después del cambio de src
                    captcha_img = driver.find_element(
                        By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage"
                    )
                    logging.info("Imagen del captcha modal recargada (src cambió)")
                except TimeoutException:
                    logging.warning("El src de la imagen no cambió en 8s, intentando de todas formas...")

            # Esperar a que la imagen tenga dimensiones reales
            WebDriverWait(driver, 10).until(lambda d: captcha_img.size['width'] > 0)

            # Esperar a que la imagen esté completamente cargada (naturalWidth > 0)
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: d.execute_script(
                        "return arguments[0].naturalWidth > 0 && arguments[0].complete;",
                        captcha_img
                    )
                )
            except TimeoutException:
                logging.warning("La imagen del captcha no terminó de cargar, esperando 2s extra...")
                time.sleep(2)

            # Guardar el src actual para comparar en el siguiente intento
            try:
                prev_src = captcha_img.get_attribute("src")
            except Exception:
                prev_src = None

            # Pequeña pausa para asegurar renderizado completo
            time.sleep(0.5)

            captcha_bytes = captcha_img.screenshot_as_png
            image = Image.open(BytesIO(captcha_bytes))
            image = image.convert('L')
            width, height = image.size
            image = image.resize((width * 2, height * 2), Image.Resampling.LANCZOS)

            text = pytesseract.image_to_string(image, config='--psm 7').strip()
            # Limpiar: solo dígitos (el captcha de Supercias siempre es numérico)
            text = re.sub(r'\D', '', text)
            logging.info(f"Texto captcha modal extraído (limpio): {text}")

            # Si el OCR no pudo extraer dígitos, no enviar captcha vacío
            if not text:
                logging.warning("OCR no pudo extraer dígitos del captcha modal, reintentando...")
                # Forzar recarga del captcha haciendo click en la imagen
                try:
                    captcha_img.click()
                    logging.info("Click en imagen del captcha para forzar recarga")
                    time.sleep(2)
                    # Actualizar prev_src para que el wait del siguiente intento funcione
                    try:
                        prev_src = driver.find_element(
                            By.CSS_SELECTOR, "#frmCaptcha\\:captchaImage"
                        ).get_attribute("src")
                    except Exception:
                        pass
                except Exception:
                    time.sleep(1)
                continue

            captcha_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "#frmCaptcha\\:captcha"))
            )
            captcha_input.clear()
            captcha_input.send_keys(text)

            boton_verificar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#frmCaptcha\\:btnPresentarContenido"))
            )
            try:
                boton_verificar.click()
            except Exception:
                driver.execute_script("arguments[0].click();", boton_verificar)
            logging.info("Captcha modal enviado")
            time.sleep(4)

        except TimeoutException:
            if intento == 0:
                logging.info("No se detectó captcha modal")
                return True
            else:
                logging.warning("Timeout durante intento de resolver captcha modal")
                
    # Verificar por última vez
    try:
        captcha_dialog = driver.find_element(By.CSS_SELECTOR, "#dlgCaptcha")
        if captcha_dialog.is_displayed():
            logging.error("No se pudo superar el captcha modal tras los reintentos")
            return False
    except NoSuchElementException:
        pass

    return True


def navigate_to_informacion_anual(driver):
    """
    Hace click en 'Información anual presentada' del menú izquierdo.
    Maneja el posible captcha que puede aparecer antes de mostrar la tabla.
    Retorna True si la tabla se cargó correctamente.
    """
    try:
        # XPATH evita el problema de los ':' en IDs con Firefox/geckodriver
        menu_item = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='frmMenu:menuInformacionAnualPresentada']"))
        )
        menu_item.click()
        logging.info("Click en Información anual presentada")
        time.sleep(3)

        # Puede aparecer captcha antes de mostrar la tabla
        captcha_resuelto = process_captcha_modal(driver)
        if not captcha_resuelto:
            logging.warning("Fallo al resolver el captcha del modal de información anual")
            return False

        # Esperar que cargue la tabla — usar XPATH para evitar problema con ':' en Firefox
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']"))
        )
        logging.info("Tabla de información anual cargada")
        return True

    except TimeoutException:
        logging.warning("No se pudo cargar la tabla de información anual")
        return False


def find_auditoria_externa_row(driver):
    """
    Busca la fila 'Auditoria Externa' (código 3.1.2) en la tabla paginada.
    Itera por todas las páginas si es necesario.
    Retorna (pdf_link, anio_fila) si lo encuentra.
    Si no lo encuentra, retorna (None, anio_mas_reciente_de_la_tabla).
    """
    page = 1
    latest_year_in_table = None

    while True:
        logging.info(f"Buscando 'Auditoria Externa' en página {page} de la tabla")

        try:
            # XPATH para evitar problema con ':' en IDs en Firefox
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']//tr"))
            )

            # Filas con data-ri son las filas de datos reales
            rows = driver.find_elements(By.XPATH, "//*[@id='frmInformacionCompanias:tblInformacionAnual_data']//tr[@data-ri]")

            # Extraer el año más reciente (primera fila de la primera página) como fallback
            if page == 1 and rows and latest_year_in_table is None:
                try:
                    cells = rows[0].find_elements(By.TAG_NAME, "td")
                    if cells:
                        anio_text = cells[0].text.strip()
                        if anio_text.isdigit():
                            latest_year_in_table = int(anio_text)
                            logging.info(f"Año más reciente detectado en la tabla (fallback): {latest_year_in_table}")
                except Exception:
                    pass

            for row in rows:
                try:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) < 5:
                        continue

                    codigo_text = cells[1].text.strip()
                    nombre_text = cells[2].text.strip()

                    if codigo_text == '3.1.2' or 'Auditoria Externa' in nombre_text:
                        logging.info(f"Encontrada fila: código={codigo_text}, nombre={nombre_text}")
                        # Extraer año AQUI antes de hacer click (evita StaleElement)
                        anio_fila = None
                        try:
                            anio_text = cells[0].text.strip()
                            anio_fila = int(anio_text) if anio_text.isdigit() else None
                            logging.info(f"Año extraído de la tabla: {anio_fila}")
                        except Exception:
                            pass
                        pdf_link = cells[4].find_element(By.CSS_SELECTOR, "a.ui-commandlink")
                        return pdf_link, anio_fila

                except NoSuchElementException:
                    continue

        except TimeoutException:
            logging.warning("Timeout esperando filas de la tabla")
            return None, latest_year_in_table

        # Verificar si hay página siguiente
        try:
            next_btn = driver.find_element(
                By.XPATH,
                "//*[@id='frmInformacionCompanias:tblInformacionAnual_paginator_bottom']//a[contains(@class,'ui-paginator-next')]"
            )
            if 'ui-state-disabled' in next_btn.get_attribute('class'):
                logging.info("No hay más páginas. Auditoría externa no encontrada.")
                return None, latest_year_in_table
            next_btn.click()
            time.sleep(2)
            page += 1
        except NoSuchElementException:
            logging.info("No se encontró paginador. Fin de búsqueda.")
            return None, latest_year_in_table


def extract_pdf_url_from_modal(driver):
    """
    Extrae la URL del PDF después de hacer click en el ícono PDF de Auditoría Externa.

    Supercias tiene dos flujos posibles:
      1. Modal normal (dlgPresentarDocumentoPdf): PDF embebido en <object data="...">
      2. Modal firma electrónica (dlgPresentarDocumentoPdfConFirmasElectronicas):
         Muestra aviso "se abrirá en pestaña nueva", con botón Aceptar que tiene
         la URL en su onclick: window.open('/consultaCompanias/tmp/doc.pdf','_blank')
         → Extraemos la URL del onclick SIN abrir la pestaña nueva (evita problemas
           con extensiones del navegador como Adobe PDF Viewer).

    Retorna la URL completa del PDF o None.
    """
    # Detectar qué modal apareció (esperar hasta 15s)
    modal_normal = None
    modal_firma = None

    try:
        # Esperar que aparezca alguno de los dos modales
        WebDriverWait(driver, 15).until(
            lambda d: (
                _is_visible(d, "//*[@id='dlgPresentarDocumentoPdf']") or
                _is_visible(d, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']")
            )
        )
    except TimeoutException:
        logging.warning("Ningún modal de PDF apareció en 15 segundos")
        return None

    # ---- Caso 1: Modal normal con <object> ----
    try:
        dlg_normal = driver.find_element(By.XPATH, "//*[@id='dlgPresentarDocumentoPdf']")
        if dlg_normal.is_displayed():
            try:
                obj_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//*[@id='panelInternoPresentarDocumentoPdf']//object[@type='application/pdf']")
                    )
                )
                pdf_data = obj_element.get_attribute("data")
                if pdf_data:
                    pdf_url = pdf_data.split('?')[0]
                    if not pdf_url.startswith('http'):
                        pdf_url = BASE_URL + pdf_url
                    logging.info(f"URL del PDF extraída (modal normal): {pdf_url}")
                    return pdf_url
            except TimeoutException:
                logging.warning("Modal normal visible pero sin <object> PDF")
    except NoSuchElementException:
        pass

    # ---- Caso 2: Modal de firma electrónica ----
    # El botón Aceptar tiene en su onclick la URL del PDF.
    # NO hacemos click (evita abrir pestaña nueva con extensión Adobe).
    # Solo leemos el atributo onclick y extraemos la URL.
    try:
        dlg_firma = driver.find_element(By.XPATH, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']")
        if dlg_firma.is_displayed():
            logging.info("Modal de firma electrónica detectado")
            boton_aceptar = dlg_firma.find_element(
                By.XPATH, ".//button[contains(@onclick,'window.open')]"
            )
            onclick = boton_aceptar.get_attribute("onclick")
            logging.info(f"onclick del botón Aceptar: {onclick[:150]}...")

            # Extraer URL del window.open('...', '_blank')
            match = re.search(
                r'window\.open\(.(/consultaCompanias/tmp/documento_[^.]+\.pdf)',
                onclick
            )
            if match:
                pdf_url = BASE_URL + match.group(1)
                logging.info(f"URL del PDF extraída (firma electrónica, sin abrir pestaña): {pdf_url}")
                # Cerrar el modal de firma electrónica sin abrirlo
                try:
                    close_btn = dlg_firma.find_element(
                        By.XPATH, ".//a[contains(@class,'ui-dialog-titlebar-close')]"
                    )
                    close_btn.click()
                    time.sleep(0.5)
                except Exception:
                    pass
                return pdf_url
    except NoSuchElementException:
        pass

    logging.warning("No se pudo extraer URL del PDF de ningún modal")
    return None


def _is_visible(driver, xpath):
    """Helper: retorna True si el elemento XPATH existe y está visible."""
    try:
        el = driver.find_element(By.XPATH, xpath)
        return el.is_displayed()
    except Exception:
        return False


def download_pdf_in_memory(driver, pdf_url):
    """
    Descarga el PDF en memoria usando fetch() desde el browser (mismo contexto de sesión).
    Esto evita errores 404 que ocurren cuando el archivo expira antes de que requests lo descargue.
    Fallback: requests con cookies de Selenium.
    No guarda nada en disco. Retorna BytesIO con el contenido del PDF.
    """
    # Método 1: fetch() desde el browser (más confiable, misma sesión activa)
    try:
        js_fetch = """
        return await fetch(arguments[0], {credentials: 'include'})
            .then(r => r.arrayBuffer())
            .then(buf => {
                let bytes = new Uint8Array(buf);
                let binary = '';
                for (let i = 0; i < bytes.byteLength; i++) {
                    binary += String.fromCharCode(bytes[i]);
                }
                return btoa(binary);
            });
        """
        b64 = driver.execute_async_script(
            """
            var url = arguments[0];
            var done = arguments[1];
            fetch(url, {credentials: 'include'})
                .then(r => r.arrayBuffer())
                .then(buf => {
                    var bytes = new Uint8Array(buf);
                    var binary = '';
                    for (var i = 0; i < bytes.byteLength; i++) {
                        binary += String.fromCharCode(bytes[i]);
                    }
                    done(btoa(binary));
                })
                .catch(e => done(null));
            """,
            pdf_url
        )
        if b64:
            import base64
            pdf_content = base64.b64decode(b64)
            if len(pdf_content) > 1000:  # PDF válido tiene más de 1KB
                logging.info(f"PDF descargado via browser fetch: {len(pdf_content)} bytes")
                return BytesIO(pdf_content)

    except Exception as e:
        logging.warning(f"fetch() browser falló: {e}, intentando con requests...")

    # Método 2: requests con cookies (fallback)
    try:
        selenium_cookies = driver.get_cookies()
        session = requests.Session()
        for cookie in selenium_cookies:
            session.cookies.set(cookie['name'], cookie['value'])

        headers = {
            'User-Agent': driver.execute_script("return navigator.userAgent;"),
            'Referer': driver.current_url,
        }

        response = session.get(pdf_url, headers=headers, timeout=30)
        response.raise_for_status()

        pdf_bytes = BytesIO(response.content)
        logging.info(f"PDF descargado via requests: {len(response.content)} bytes")
        return pdf_bytes

    except Exception as e:
        logging.error(f"Error descargando PDF: {e}")
        return None


def _extract_firma_auditora_heuristic(pdf_bytes):
    """
    Extrae firma auditora y socio firmante del PDF de Auditoria Externa de Supercias (método heurístico).

    Retorna string combinado: "PKFECUADOR & CO. C.L. / Manuel Garcia Andrade (SC-RNAE-002)"
    O solo nombre si no hay empresa separada: "Ing. Paola Zamora C (SC-RNAE-1562)"

    Maneja:
      - PDF digital (pdfplumber): firmas empresa + socio con firma electronica
      - PDF escaneado (OCR): auditor persona natural con sello en pie de pagina
    """

    TITULOS_A_IGNORAR = [
        'dictamen de los auditores', 'informe de auditor', 'informe del auditor',
        'auditoria externa', 'estados financieros', 'parte i', 'parte ii',
        'externoindependiente', 'externo independiente', 'informe sobre',
        'cuenca, marzo', 'cuenca, ecuador', 'samborondon, ecuador',
        'cumplimiento tributario', 'de cumplimiento', 'responsabilidades',
        'base de la opinion', 'incertidumbre material', 'opinion:',
        'hemos auditado', 'hemos llevado', 'a los accionistas',
        # Variantes OCR sin primera letra (común en PDFs escaneados)
        'nforme de los auditores', 'nforme de auditores',
        'nforme de los auditores independientes',
        'nforme de los auditores externos',
        'nforme del auditor', 'ndice. paginas',
        # Contenido legal/regulatorio que no es nombre de firma
        'obligaciones tributarias', 'regimen tributario',
        'registro nacional de auditores', 'informacion financiera suplementaria',
        'resolucion n', 'validar unicamente', 'firma electronica',
        'control que identifique', 'en el transcurso de la auditoria',
        'en la realizacion de la auditoria',
        'emitido sentencia', 'esta ley no dispone',
    ]
    PREFIJOS_PROF = ['ing.', 'dr.', 'dra.', 'lcdo.', 'lcda.', 'cpa ', 'eco.', 'mba.', 'abg.']
    # Patrones de substring para detectar firmas de empresa
    # Nota: 'ey' se maneja con regex word-boundary aparte (ver _es_patron_empresa)
    PATRONES_EMPRESA_SUBSTRING = [
        'pkf', 'deloitte', 'kpmg', 'ernst & young', 'grant thornton',
        'bdo ', 'moore ', 'crowe', 'mazars', 'nexia', 'baker tilly',
        '& co.', '& co. c.l.', '& asociados', 'auditores y consultores',
        'firma auditora', 'cia. ltda', 'cia.ltda',
        'audit ', 'consulting',
    ]
    # Patrones que requieren word-boundary (evitar falsos positivos con "Ley", etc.)
    _PATRONES_EMPRESA_REGEX = [
        re.compile(r'\bey\b', re.IGNORECASE),
    ]

    def _es_patron_empresa(texto_lower):
        """Verifica si el texto contiene algún patrón de empresa (substring o regex)."""
        if any(p in texto_lower for p in PATRONES_EMPRESA_SUBSTRING):
            return True
        if any(rx.search(texto_lower) for rx in _PATRONES_EMPRESA_REGEX):
            return True
        return False

    def es_titulo(linea):
        return any(t in linea.lower() for t in TITULOS_A_IGNORAR)

    def es_contenido_legal(linea):
        """Detecta líneas que son contenido legal/regulatorio, no nombres de firmas."""
        l = linea.lower()
        indicadores = [
            'ley de', 'ley no', 'esta ley', 'reglament', 'sentencia',
            'obligacion', 'tributari', 'regimen', 'resolucion',
            'norma', 'disposicion', 'articulo', 'decreto',
            'suplementaria', 'requerida por', 'establecidas por',
            'identificar en', 'identifique en', 'transcurso de',
            'realizacion de la auditoria', 'validar unicamente',
        ]
        return any(ind in l for ind in indicadores)

    def es_ruido_ocr(linea):
        palabras = linea.split()
        if not palabras: return True
        cortas = sum(1 for p in palabras if len(re.sub(r'[^a-zA-Z]','',p)) <= 2)
        return len(palabras) > 5 and cortas / len(palabras) > 0.6

    def limpiar_ocr(linea):
        linea = re.sub(r'^[|\[\]!l1I]+\s*', '', linea).strip()
        return re.sub(r'[|_]{2,}', '', linea).strip()

    def normalizar_nombre(nombre):
        nombre = re.sub(r'^[nN][gG]\.', 'Ing.', nombre)
        nombre = re.sub(r'^\([nN][gG][,\.]', 'Ing.', nombre)
        nombre = re.sub(r'^\|\s*aola', 'Paola', nombre)
        nombre = re.sub(r'\s*\|.*$', '', nombre)
        return nombre.rstrip('.|,').strip()

    def extraer_rnae_numero(texto):
        """Extrae el numero de registro RNAE de un texto. Ej: SC-RNAE-002"""
        match = re.search(r'SC-?RN\s?AE[-\s]*(\d+)', texto, re.IGNORECASE)
        if match:
            return f"SC-RNAE-{match.group(1)}"
        return None

    def puntaje_candidato(nombre):
        nombre_norm = normalizar_nombre(nombre)
        nombre_l = nombre_norm.lower()
        puntaje = 0
        
        # Priorización de coincidencia con Excel
        if AUDITORES_SET:
            if _normalizar_texto(nombre) in AUDITORES_SET:
                puntaje += 100  # Máxima prioridad

        if _es_patron_empresa(nombre_l): puntaje += 25
        if any(nombre_l.startswith(p) for p in PREFIJOS_PROF): puntaje += 20
        if 5 <= len(nombre_norm) <= 80: puntaje += 5
        if len(nombre_norm.split()) >= 2: puntaje += 5
        raros = sum(1 for c in nombre_norm if not (c.isalpha() or c in ' .-,°()&'))
        puntaje -= raros * 3
        palabras_1 = sum(1 for p in nombre_norm.split()
                        if len(re.sub(r'[^a-zA-Z]','',p)) == 1)
        puntaje -= palabras_1 * 5
        return puntaje, nombre_norm

    def buscar_en_lineas(lineas, solo_primeras_lineas=False):
        """
        Retorna dict con:
          candidatos_empresa: [(nombre, pagina)]  -- firmas empresa
          candidatos_persona: [(nombre, pagina)]  -- personas naturales
          socio:    nombre del socio firmante (extraido junto al RNAE)
          rnae_num: numero SC-RNAE-XXX
        """
        resultado = {'empresa': [], 'persona': [], 'socio': None, 'rnae': None}
        rango = range(min(5, len(lineas))) if solo_primeras_lineas else range(len(lineas))

        for i in rango:
            linea = lineas[i]
            lc = limpiar_ocr(linea.strip())
            lc_norm = normalizar_nombre(lc)
            lc_l = lc_norm.lower()

            if es_titulo(lc) or es_ruido_ocr(lc):
                continue

            # --- NUEVA LÓGICA CON EXCEL ---
            if AUDITORES_SET:
                match_excel = buscar_auditor_en_linea(linea, AUDITORES_SET, MAPA_ORIGINALES_AUD)
                if match_excel:
                    logging.info(f"Coincidencia de Excel encontrada para la línea '{linea}': '{match_excel}'")
                    match_l = match_excel.lower()
                    if _es_patron_empresa(match_l):
                        resultado['empresa'].append(match_excel)
                    else:
                        resultado['persona'].append(match_excel)

            # Firma empresa en linea propia
            if _es_patron_empresa(lc_l):
                if (3 < len(lc) <= 80 and len(lc.split()) <= 8
                        and not es_titulo(lc) and not es_contenido_legal(lc)):
                    resultado['empresa'].append(lc_norm)

            # RNAE en esta linea
            if ('rnae' in lc_l or 'sc-rnae' in lc_l or
                    'registro no.' in lc_l or 'registro no. sc' in lc_l or
                    ('auditor scv' in lc_l and 'rnae' in lc_l)):

                # Extraer numero RNAE
                rnae = extraer_rnae_numero(lc)
                if rnae and not resultado['rnae']:
                    resultado['rnae'] = rnae

                # Nombre del socio en la MISMA linea (despues del RNAE)
                match = re.search(
                    r'(?:SC-?RN\s+AE|SC-?RNAE)-?\s*\d+\s+(.+)',
                    lc, re.IGNORECASE
                )
                if match:
                    socio_inline = match.group(1).strip().rstrip('.,')
                    if len(socio_inline) > 5 and len(socio_inline.split()) >= 2:
                        if AUDITORES_SET:
                            match_excel = buscar_auditor_en_linea(socio_inline, AUDITORES_SET, MAPA_ORIGINALES_AUD)
                            if match_excel:
                                socio_inline = match_excel
                        resultado['socio'] = socio_inline

                # Nombre del socio/auditor en lineas ANTERIORES
                if not resultado['socio']:
                    for j in range(max(0, i-4), i):
                        c = limpiar_ocr(lineas[j].strip())
                        c_norm = normalizar_nombre(c)
                        if (len(c_norm) > 6 and not es_titulo(c_norm)
                                and not re.match(r'^[\d\s,\.°]+$', c_norm)
                                and not es_ruido_ocr(c_norm)
                                and 'cuenca' not in c_norm.lower()
                                and 'ecuador' not in c_norm.lower()
                                and 'samborondon' not in c_norm.lower()
                                and 'registro' not in c_norm.lower()
                                and len(c_norm.split()) >= 2):
                            if AUDITORES_SET:
                                match_excel = buscar_auditor_en_linea(c_norm, AUDITORES_SET, MAPA_ORIGINALES_AUD)
                                if match_excel:
                                    c_norm = match_excel
                            resultado['persona'].append(c_norm)
                            resultado['socio'] = resultado['socio'] or c_norm

            # Titulo profesional cerca de auditor/RNAE
            if any(lc_l.startswith(p) for p in PREFIJOS_PROF):
                ctx = ' '.join(lineas[max(0,i-2):min(len(lineas),i+5)]).lower()
                if ('auditor' in ctx or 'rnae' in ctx or 'socio' in ctx or
                        'scvc' in ctx or 'scvg' in ctx or 'pkf' in ctx):
                    if len(lc_norm) > 8 and len(lc_norm.split()) >= 3:
                        resultado['persona'].append(lc_norm)

        return resultado

    def ocr_zona(doc_fitz, num_pagina, desde=0.60):
        try:
            import fitz
            page = doc_fitz[num_pagina]
            mat = fitz.Matrix(3, 3)
            pix = page.get_pixmap(matrix=mat)
            image = Image.open(BytesIO(pix.tobytes("png"))).convert('L')
            w, h = image.size
            zona = image.crop((0, int(h * desde), w, h))
            return pytesseract.image_to_string(zona, config='--psm 6')
        except Exception as e:
            logging.warning(f"Error OCR zona pag {num_pagina+1}: {e}")
            return ""

    # Acumuladores globales
    todas_empresas = []
    todos_socios = []
    todos_rnae = []
    todos_personas = []

    # ---- INTENTO 1: pdfplumber en paginas (texto digital) ----
    try:
        pdf_bytes.seek(0)
        with pdfplumber.open(pdf_bytes) as pdf:
            total_paginas = len(pdf.pages)
            logging.info(f"PDF tiene {total_paginas} paginas")

            # Analizar primeras 5 páginas y últimas 3 páginas
            paginas_set = set()
            for p in range(min(5, total_paginas)):
                paginas_set.add(p)
            for p in range(max(0, total_paginas - 3), total_paginas):
                paginas_set.add(p)
            paginas_texto = sorted(list(paginas_set))

            for num in paginas_texto:
                texto = pdf.pages[num].extract_text() or ""
                if not texto.strip(): continue
                lineas = [l.strip() for l in texto.split('\n') if l.strip()]
                solo_inicio = num < 2
                r = buscar_en_lineas(lineas, solo_primeras_lineas=solo_inicio)
                if not r['empresa'] and not r['socio'] and solo_inicio:
                    r = buscar_en_lineas(lineas)
                todas_empresas.extend([(e, num+1) for e in r['empresa']])
                todos_personas.extend([(p, num+1) for p in r['persona']])
                if r['socio']: todos_socios.append(r['socio'])
                if r['rnae']: todos_rnae.append(r['rnae'])

    except Exception as e:
        logging.warning(f"Error pdfplumber: {e}")

    # ---- INTENTO 2: OCR zona inferior paginas (PDF escaneado) ----
    if not todas_empresas and not todos_personas:
        try:
            import fitz
            pdf_bytes.seek(0)
            doc_ocr = fitz.open(stream=pdf_bytes.read(), filetype="pdf")
            total_paginas = doc_ocr.page_count
            
            # Priorizar páginas en orden: últimas 3, intermedias (4-7), y primeras (1-3)
            paginas_priorizadas = []
            
            # 1. Las últimas 3 páginas (donde están las firmas del informe completo)
            ultimas = [p for p in range(max(0, total_paginas - 3), total_paginas)]
            for p in ultimas:
                if p not in paginas_priorizadas:
                    paginas_priorizadas.append(p)
            
            # 2. Páginas 4 a 7 (donde suele estar la opinión en informes cortos)
            intermedias = [p for p in range(4, min(8, total_paginas))]
            for p in intermedias:
                if p not in paginas_priorizadas:
                    paginas_priorizadas.append(p)
                    
            # 3. Páginas 1 a 3 (fallback, omitiendo la portada 0 si es posible)
            start_p = 1 if total_paginas > 1 else 0
            primeras = [p for p in range(start_p, min(4, total_paginas))]
            for p in primeras:
                if p not in paginas_priorizadas:
                    paginas_priorizadas.append(p)
            
            # Limitar a máximo 8 páginas totales para evitar lentitud extrema en OCR
            paginas = paginas_priorizadas[:8]
            logging.info(f"Páginas seleccionadas para OCR (priorizadas): {paginas}")

            for num in paginas:
                texto = ocr_zona(doc_ocr, num, desde=0.60)
                lineas = [l.strip() for l in texto.split('\n') if l.strip()]
                r = buscar_en_lineas(lineas)
                todas_empresas.extend([(e, num+1) for e in r['empresa']])
                todos_personas.extend([(p, num+1) for p in r['persona']])
                if r['socio']: todos_socios.append(r['socio'])
                if r['rnae']: todos_rnae.append(r['rnae'])
            doc_ocr.close()
        except ImportError:
            logging.warning("PyMuPDF no disponible")
        except Exception as e:
            logging.warning(f"Error OCR: {e}")

    # ---- Elegir mejor empresa y mejor persona por puntaje ----
    mejor_empresa = None
    mejor_persona = None
    mejor_rnae = todos_rnae[0] if todos_rnae else None

    if todas_empresas:
        evaluados = [(puntaje_candidato(n)[0], puntaje_candidato(n)[1], pg)
                     for n, pg in todas_empresas]
        evaluados.sort(reverse=True)
        mejor_empresa = evaluados[0][1]
        logging.info(f"Empresa auditora: '{mejor_empresa}'")

    if todos_personas:
        evaluados = [(puntaje_candidato(n)[0], puntaje_candidato(n)[1], 0)
                     for n, _ in todos_personas]
        evaluados.sort(reverse=True)
        mejor_persona = evaluados[0][1]
        logging.info(f"Persona auditora: '{mejor_persona}'")

    # Elegir el mejor socio firmante comparando puntajes
    socio_final = None
    if todos_socios:
        # Evaluar todos los socios candidatos y elegir el de mayor puntaje
        socios_eval = [(puntaje_candidato(s)[0], puntaje_candidato(s)[1]) for s in todos_socios]
        socios_eval.sort(reverse=True)
        socio_final = socios_eval[0][1]
        logging.info(f"Socio firmante candidatos: {socios_eval}")
        logging.info(f"Socio firmante seleccionado: '{socio_final}'")

    # Si la persona natural tiene mejor puntaje que el socio extraido del inline, usarla
    if mejor_persona and socio_final:
        p_socio, _ = puntaje_candidato(socio_final)
        p_persona, _ = puntaje_candidato(mejor_persona)
        if p_persona > p_socio:
            logging.info(f"Usando persona ({p_persona}) en vez de socio inline ({p_socio})")
            socio_final = mejor_persona
    elif mejor_persona and not socio_final:
        socio_final = mejor_persona

    # ---- Construir el resultado final ----
    # Casos:
    # 1. Empresa + socio: "PKFECUADOR & CO. C.L. / Manuel Garcia Andrade (SC-RNAE-002)"
    # 2. Solo empresa:    "PKFECUADOR & CO. C.L."
    # 3. Persona natural: "Ing. Paola Zamora C (SC-RNAE-1562)"
    # 4. Solo RNAE+socio: "Manuel Garcia Andrade (SC-RNAE-002)"

    partes = []

    if mejor_empresa:
        partes.append(mejor_empresa)

    nombre_auditor = socio_final or mejor_persona
    if nombre_auditor:
        if mejor_rnae:
            partes.append(f"{nombre_auditor} ({mejor_rnae})")
        else:
            partes.append(nombre_auditor)
    elif mejor_rnae and not mejor_empresa:
        partes.append(f"RNAE {mejor_rnae}")

    if partes:
        resultado_final = ' / '.join(partes)
        logging.info(f"Resultado final: '{resultado_final}'")
        return resultado_final

    logging.info("No se encontro firma auditora en el PDF")
    return None


def extract_firma_auditora_from_pdf(pdf_bytes):
    """
    Extrae la firma auditora, el auditor o socio firmante utilizando un modelo LLM local (Ollama).
    Si Ollama no está disponible o falla, delega automáticamente a la heurística tradicional.
    """
    if not USE_OLLAMA or ollama is None:
        logging.info("[Ollama] Desactivado o librería no disponible. Usando extractor heurístico.")
        return _extract_firma_auditora_heuristic(pdf_bytes)

    logging.info(f"[Ollama] Iniciando extracción con modelo: '{OLLAMA_MODEL}' en '{OLLAMA_HOST}'...")
    texto_relevante = ""
    
    # ---- 1. EXTRACCIÓN DE TEXTO DIGITAL (pdfplumber) ----
    try:
        pdf_bytes.seek(0)
        with pdfplumber.open(pdf_bytes) as pdf:
            total_paginas = len(pdf.pages)
            logging.info(f"[Ollama Context] El PDF tiene {total_paginas} páginas.")

            # Seleccionar páginas clave: primeras 3 y últimas 2
            paginas_clave = set()
            for p in range(min(3, total_paginas)):
                paginas_clave.add(p)
            for p in range(max(0, total_paginas - 2), total_paginas):
                paginas_clave.add(p)
            
            for num in sorted(list(paginas_clave)):
                texto_pag = pdf.pages[num].extract_text() or ""
                if texto_pag.strip():
                    texto_relevante += f"\n--- TEXTO PÁGINA {num+1} ---\n{texto_pag}"
    except Exception as e:
        logging.warning(f"[Ollama] Error al extraer con pdfplumber: {e}")

    # ---- 2. FALLBACK A OCR SI NO SE DETECTÓ TEXTO DIGITAL (PDF Escaneado) ----
    if not texto_relevante.strip():
        logging.info("[Ollama Context] PDF parece escaneado. Aplicando OCR a páginas clave...")
        try:
            import fitz
            pdf_bytes.seek(0)
            doc_ocr = fitz.open(stream=pdf_bytes.read(), filetype="pdf")
            total_paginas = doc_ocr.page_count
            
            paginas_clave = set()
            for p in range(min(3, total_paginas)):
                paginas_clave.add(p)
            for p in range(max(0, total_paginas - 2), total_paginas):
                paginas_clave.add(p)

            for num in sorted(list(paginas_clave)):
                try:
                    page = doc_ocr[num]
                    mat = fitz.Matrix(2.5, 2.5) # Balance óptimo velocidad/calidad
                    pix = page.get_pixmap(matrix=mat)
                    image = Image.open(BytesIO(pix.tobytes("png"))).convert('L')
                    texto_pag = pytesseract.image_to_string(image, config='--psm 3')
                    if texto_pag.strip():
                        texto_relevante += f"\n--- TEXTO OCR PÁGINA {num+1} ---\n{texto_pag}"
                except Exception as ex_pag:
                    logging.warning(f"[Ollama] Error OCR en página {num+1}: {ex_pag}")
            doc_ocr.close()
        except Exception as e:
            logging.error(f"[Ollama] Error crítico en fallback OCR: {e}")

    if not texto_relevante.strip():
        logging.warning("[Ollama] No se pudo extraer texto de ninguna forma del PDF. Usando fallback heurístico.")
        return _extract_firma_auditora_heuristic(pdf_bytes)

    # ---- 3. PROCESAMIENTO ESTRUCTURADO CON OLLAMA ----
    logging.info("[Ollama] Enviando fragmentos de texto seleccionados a Ollama...")
    
    prompt_sistema = """
    Eres un asistente de IA experto en auditoría legal y corporativa en Ecuador. 
    Tu única tarea es leer los fragmentos de texto de un informe de la Superintendencia de Compañías (SuperCias) y extraer la Identidad del Auditor Externo (Firma Auditora o Persona Natural).

    Reglas estrictas para evitar FALSOS POSITIVOS:
    1. Responde ÚNICAMENTE con un objeto JSON válido que contenga las llaves: "firma_auditora", "socio_firmante", "rnae".
    2. En "firma_auditora", coloca el nombre de la compañía/firma (ej: "BDO Ecuador", "PriceWaterhouseCoopers", "Moore", etc.). Si es una persona natural actuando de forma independiente, pon null.
    3. En "socio_firmante", coloca exclusivamente el nombre del ser humano que firma el dictamen (ej: "Ing. Juan Pérez", "CPA María Rodríguez").
    4. En "rnae", coloca el número del Registro Nacional de Auditores Externos si aparece (ej: "SC-RNAE-002" o "RNAE-123" o "123").
    5. CRÍTICO: No inventes datos. No confundas al Gerente General, Representante Legal, Comisario o Presidente de la empresa auditada con el Auditor. Si no estás 100% seguro de un campo, coloca null en ese campo.
    """

    prompt_usuario = f"A continuación se presenta el texto extraído del informe. Analízalo rigurosamente:\n\n{texto_relevante}"

    try:
        # Configurar host del cliente si es necesario
        client = ollama.Client(host=OLLAMA_HOST)
        response = client.generate(
            model=OLLAMA_MODEL,
            system=prompt_sistema,
            prompt=prompt_usuario,
            format='json',
            options={
                "temperature": 0.0  # Determinismo absoluto
            }
        )

        # Parsear la respuesta de Ollama
        respuesta_cruda = response['response']
        datos_json = json.loads(respuesta_cruda)
        logging.info(f"[Ollama Response] JSON Recibido: {datos_json}")

        # ---- 4. CRUCE Y NORMALIZACIÓN CON EXCEL DE AUDITORES ----
        firma_auditora = datos_json.get("firma_auditora")
        socio_firmante = datos_json.get("socio_firmante")
        rnae = datos_json.get("rnae")

        # Limpiar strings
        if firma_auditora: firma_auditora = firma_auditora.strip()
        if socio_firmante: socio_firmante = socio_firmante.strip()
        if rnae: rnae = rnae.strip()

        # Si el RNAE es solo dígitos, formatearlo como "SC-RNAE-XXX"
        if rnae:
            rnae_digits = re.search(r'\d+', rnae)
            if rnae_digits:
                rnae = f"SC-RNAE-{int(rnae_digits.group()):03d}"

        # Realizar cruce de coincidencia difusa con la lista cargada desde Excel si está disponible
        if AUDITORES_SET:
            if firma_auditora:
                match_firma = buscar_auditor_en_linea(firma_auditora, AUDITORES_SET, MAPA_ORIGINALES_AUD)
                if match_firma:
                    logging.info(f"[Ollama Normalization] Firma '{firma_auditora}' -> '{match_firma}' (Excel)")
                    firma_auditora = match_firma
            if socio_firmante:
                match_socio = buscar_auditor_en_linea(socio_firmante, AUDITORES_SET, MAPA_ORIGINALES_AUD)
                if match_socio:
                    logging.info(f"[Ollama Normalization] Socio '{socio_firmante}' -> '{match_socio}' (Excel)")
                    socio_firmante = match_socio

        # ---- 5. RECONSTRUCCIÓN DEL FORMATO REQUERIDO POR LA BASE DE DATOS ----
        partes = []
        if firma_auditora:
            partes.append(firma_auditora)
        
        if socio_firmante:
            if rnae:
                partes.append(f"{socio_firmante} ({rnae})")
            else:
                partes.append(socio_firmante)
        elif rnae and not firma_auditora:
            partes.append(f"RNAE {rnae}")

        if partes:
            resultado_final = ' / '.join(partes)
            logging.info(f"[Ollama] Resultado final estructurado: '{resultado_final}'")
            return resultado_final
        
        logging.info("[Ollama] La extracción devolvió campos vacíos. Usando fallback heurístico.")
        return _extract_firma_auditora_heuristic(pdf_bytes)

    except Exception as e:
        logging.error(f"[Ollama] Falló el procesamiento con Ollama ({e}). Usando fallback heurístico.")
        try:
            return _extract_firma_auditora_heuristic(pdf_bytes)
        except Exception as ex_fallback:
            logging.error(f"[Ollama] Error crítico en el fallback heurístico: {ex_fallback}")
            return None


def close_pdf_modal(driver):
    """Cierra el modal del PDF si está abierto."""
    try:
        # Intentar cerrar dlgPresentarDocumentoPdf
        close_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='dlgPresentarDocumentoPdf']//a[contains(@class,'ui-dialog-titlebar-close')]"))
        )
        close_btn.click()
        time.sleep(1)
        logging.info("Modal PDF cerrado")
    except TimeoutException:
        pass

    try:
        # Intentar cerrar dlgPresentarDocumentoPdfConFirmasElectronicas
        close_btn = driver.find_element(
            By.XPATH, "//*[@id='dlgPresentarDocumentoPdfConFirmasElectronicas']//a[contains(@class,'ui-dialog-titlebar-close')]"
        )
        if close_btn.is_displayed():
            close_btn.click()
            time.sleep(1)
    except NoSuchElementException:
        pass


def process_auditoria_externa(driver, expediente):
    """
    Orquesta todo el flujo de obtención de la firma auditora:
    1. Navegar a Información Anual Presentada
    2. Buscar fila Auditoría Externa (código 3.1.2) en la tabla paginada
    3. Click en el PDF → resolver captcha si aparece
    4. Extraer URL del PDF del modal
    5. Descargar PDF en memoria
    6. Extraer firma auditora del PDF
    
    Retorna dict con {'firma': str, 'anio': int|None} o None si falla completamente la navegación.
    """
    resultado = {'firma': None, 'anio': None}

    try:
        # Paso 1: Navegar a Información Anual
        if not navigate_to_informacion_anual(driver):
            logging.warning(f"[{expediente}] No se pudo navegar a Información Anual")
            return None

        # Paso 2: Buscar la fila de Auditoría Externa
        pdf_link, anio_fila = find_auditoria_externa_row(driver)
        if not pdf_link:
            logging.info(f"[{expediente}] No tiene documento de Auditoría Externa. Registrando S/A con año NULL.")
            resultado['firma'] = 'S/A'
            resultado['anio'] = None
            return resultado

        resultado['anio'] = anio_fila
        logging.info(f"[{expediente}] Año auditoría: {resultado['anio']}")

        # Paso 3: Click en el PDF
        logging.info(f"[{expediente}] Haciendo click en PDF de Auditoría Externa")
        pdf_link.click()
        time.sleep(3)

        # Resolver captcha si aparece antes del modal
        captcha_resuelto = process_captcha_modal(driver)
        if not captcha_resuelto:
            logging.warning(f"[{expediente}] No se pudo superar el captcha modal")
            resultado['firma'] = 'N/A'
            return resultado

        # Paso 4: Extraer URL del PDF desde el modal
        pdf_url = extract_pdf_url_from_modal(driver)
        if not pdf_url:
            logging.warning(f"[{expediente}] No se pudo obtener URL del PDF")
            close_pdf_modal(driver)
            resultado['firma'] = 'N/A'
            return resultado

        # Paso 5: Descargar PDF en memoria (con cookies de sesión)
        pdf_bytes = download_pdf_in_memory(driver, pdf_url)
        if not pdf_bytes:
            logging.warning(f"[{expediente}] Error descargando el PDF")
            close_pdf_modal(driver)
            resultado['firma'] = 'N/A'
            return resultado

        # Paso 6: Extraer firma auditora del PDF
        firma = extract_firma_auditora_from_pdf(pdf_bytes)
        resultado['firma'] = firma if (firma and firma.strip()) else 'N/A'

        logging.info(f"[{expediente}] Resultado auditoría: firma={resultado['firma']}, anio={resultado['anio']}")

        # Cerrar el modal
        close_pdf_modal(driver)

        return resultado

    except Exception as e:
        logging.error(f"[{expediente}] Error en process_auditoria_externa: {e}")
        try:
            close_pdf_modal(driver)
        except Exception:
            pass
        # Si falló por alguna excepción inesperada pero ya logramos extraer el año
        if resultado.get('anio'):
            resultado['firma'] = 'N/A'
            return resultado
        return None


def insertar_auditoria_en_bd(cursor, conn, id_compania, firma, anio):
    """
    Inserta el registro de auditoría en la tabla com_companias_auditorias_previas.
    """
    if not firma:
        logging.info(f"Firma vacía, no se inserta en BD")
        return False

    try:
        # Verificar si ya existe un registro para esa compañía y año
        if anio is None:
            check_query = """
                SELECT id FROM crm.com_companias_auditorias_previas 
                WHERE id_compania = %s AND anio_auditoria IS NULL;
            """
            cursor.execute(check_query, (id_compania,))
        else:
            check_query = """
                SELECT id FROM crm.com_companias_auditorias_previas 
                WHERE id_compania = %s AND anio_auditoria = %s;
            """
            cursor.execute(check_query, (id_compania, anio))
        
        existente = cursor.fetchone()

        if existente:
            # Actualizar si ya existe
            if anio is None:
                update_query = """
                    UPDATE crm.com_companias_auditorias_previas 
                    SET nombre_firma = %s, fecha_creacion = now()
                    WHERE id_compania = %s AND anio_auditoria IS NULL;
                """
                cursor.execute(update_query, (firma, id_compania))
            else:
                update_query = """
                    UPDATE crm.com_companias_auditorias_previas 
                    SET nombre_firma = %s, fecha_creacion = now()
                    WHERE id_compania = %s AND anio_auditoria = %s;
                """
                cursor.execute(update_query, (firma, id_compania, anio))
            logging.info(f"Registro de auditoría actualizado: id_compania={id_compania}, firma={firma}, anio={anio}")
        else:
            # Insertar nuevo
            insert_query = """
                INSERT INTO crm.com_companias_auditorias_previas 
                    (id_compania, nombre_firma, anio_auditoria, fecha_creacion)
                VALUES (%s, %s, %s, now());
            """
            cursor.execute(insert_query, (id_compania, firma, anio))
            logging.info(f"Registro de auditoría insertado: id_compania={id_compania}, firma={firma}, anio={anio}")

        conn.commit()
        return True

    except Exception as e:
        logging.error(f"Error insertando auditoría en BD: {e}")
        conn.rollback()
        return False


# ==========================================
# LÓGICA PRINCIPAL - ORIGINAL
# ==========================================
def search_company(compania, driver, new_connection):
    """Realiza la búsqueda de una compañía para luego extraer sus datos."""
    success, auto_complete_count = _buscar_y_cargar_compania(driver, compania)
    if not success:
        return None

    correos = []
    try:
        correos = extract_company_data(driver, auto_complete_count)
    except StaleElementReferenceException:
        pass
    finally:
        elements = driver.find_elements(By.CSS_SELECTOR, ".ui-messages-error-summary")
        if elements:
            refresh_session(driver)
        else:
            try:
                nueva_busqueda = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "#frmNuevaConsulta>button"))
                )
                nueva_busqueda.click()
            except Exception:
                pass

    return correos


def process_database_records(error_file, processed_file, driver, new_connection):
    """
    Procesa los registros desde PostgreSQL.
    Para empresas con estado_proceso = 'T': ya tienen datos generales,
    solo extraemos la firma auditora del PDF de Auditoria Externa.
    """
    l_i = 1

    try:
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD
        )
        cursor = conn.cursor()

        #Procesar pendientes (T) + los que quedaron en F sin firma registrada
        cursor.execute("""
            SELECT expediente, id
            FROM crm.adm_companias
            WHERE estado_proceso = 'T'
              AND fecha_actualizacion = %s
              AND activos >= %s;
        """, (DB_FECHA_ACTUALIZACION, DB_MIN_ACTIVOS))
        records = cursor.fetchall()

    except Exception as e:
        logging.error(f"Error conectando a la base de datos: {e}")
        return

    try:
        with open(error_file, 'a', newline='') as csvfile_error, \
             open(processed_file, 'a', newline='') as csvfile_processed:

            error_writer = csv.writer(csvfile_error, delimiter=';')
            processed_writer = csv.writer(csvfile_processed, delimiter=';')

            for record in records:
                expediente = record[0]
                id_compania = record[1]
                row = [expediente]

                logging.info(f"Row: {l_i}")
                logging.info(f"Company: {expediente} (id={id_compania})")

                success = False
                for intento in range(3):
                    try:
                        # Navegar directo a la compania y extraer auditoria
                        logging.info(f"[{expediente}] Navegando para extraccion de Auditoria Externa...")
                        _navegar_a_compania(driver, expediente)
                        auditoria = process_auditoria_externa(driver, expediente)

                        # auditoria == None significa fallo de navegación/captcha → reintentar
                        if auditoria is None:
                            logging.warning(f"[{expediente}] Fallo de navegación/captcha (intento {intento+1}/3), reintentando...")
                            time.sleep(2)
                            continue

                        # Guardar auditoria en BD (firma puede ser S/A o N/A, ambos son resultados válidos)
                        if auditoria.get('firma'):
                            insertar_auditoria_en_bd(
                                cursor, conn,
                                id_compania=id_compania,
                                firma=auditoria['firma'],
                                anio=auditoria.get('anio')
                            )
                        else:
                            logging.info(f"[{expediente}] Sin firma auditora para registrar")

                        processed_writer.writerow(row)
                        success = True

                        # Actualizar estado_proceso a 'P' (Procesado)
                        try:
                            cursor.execute("""
                                UPDATE crm.adm_companias
                                SET estado_proceso = 'P'
                                WHERE id = %s;
                            """, (id_compania,))
                            conn.commit()
                            logging.info(f"[{expediente}] estado_proceso actualizado a 'P'")
                        except Exception as e_upd:
                            logging.warning(f"[{expediente}] No se pudo actualizar estado_proceso: {e_upd}")
                            conn.rollback()

                        break

                    except Exception as e:
                        logging.info(f"Error procesando {expediente} (intento {intento+1}): {e}")
                        time.sleep(2)

                if not success:
                    logging.warning(f"[{expediente}] No se pudo procesar tras 3 intentos. Registrando en archivo de errores.")
                    error_writer.writerow(row)

                l_i += 1

    finally:
        if cursor:
            cursor.close()
            logging.info("Cursor de BD cerrado.")
        if conn:
            conn.close()
            logging.info("Conexion a BD cerrada correctamente.")


def _navegar_a_compania(driver, expediente):
    """
    Navega a la página de información de una compañía por su expediente.
    Se usa para acceder al menú después de que search_company terminó.
    Reutiliza el flujo de búsqueda sin volver a actualizar la BD.
    """
    success, _ = _buscar_y_cargar_compania(driver, expediente)
    if not success:
        raise Exception(f"No se pudo cargar la compañía {expediente} (falló búsqueda o captcha)")


# ==========================================
# EJECUCIÓN
# ==========================================
def main():
    driver = None
    new_connection = False
    start_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        logging.info("Inicio del script.")

        options = Options()
        options.binary_location = FIREFOX_BINARY_PATH
        if FIREFOX_HEADLESS:
            options.add_argument('--headless')
        driver = webdriver.Firefox(options=options)

        try:
            process_database_records(ERROR_FILE, PROCESSED_FILE, driver, new_connection)
        except KeyboardInterrupt:
            logging.info("Proceso interrumpido por el usuario.")

        end_time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{start_time_str} - {end_time_str}] Fin del script.")

    finally:
        if driver is not None:
            driver.quit()


if __name__ == "__main__":
    main()